"""Router REST do Assistente.

Endpoints para:
  - Importar clientes das planilhas
  - Registrar eventos/datas importantes
  - Agendar reuniões
  - Consultar agenda e próximos eventos
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import structlog
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from src.assistente.adapters.outbound.importers.planilha_clientes_importer import (
    importar_clientes,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.cliente_assessor_repository import (
    SqlAlchemyClienteAssessorRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.evento_repository import (
    SqlAlchemyEventoRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.reuniao_repository import (
    SqlAlchemyReuniaoRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.renda_fixa_repository import (
    SqlAlchemyRendaFixaRepository,
)
from src.assistente.application.commands.agendar_reuniao import AgendarReuniaoCommand
from src.assistente.application.commands.registrar_evento import RegistrarEventoCommand
from src.assistente.application.commands.importar_diversificador import ImportarDiversificadorCommand
from src.assistente.application.handlers.agendar_reuniao_handler import AgendarReuniaoHandler
from src.assistente.application.handlers.registrar_evento_handler import RegistrarEventoHandler
from src.assistente.application.handlers.importar_diversificador_handler import ImportarDiversificadorHandler
from src.assistente.adapters.outbound.persistence.sqlalchemy.posicoes_rv_repository import (
    SqlAlchemyPosicaoRVRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.posicoes_fundos_repository import (
    SqlAlchemyPosicaoFundoRepository,
    SqlAlchemyPosicaoPrevRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.lead_repository import (
    SqlAlchemyLeadRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.anotacao_repository import (
    SqlAlchemyAnotacaoRepository,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.historico_imports_repository import (
    SqlAlchemyHistoricoImportsRepository,
)
from src.assistente.application.handlers.importar_diversificador_completo_handler import (
    ImportarDiversificadorCompletoHandler,
)
from src.assistente.adapters.outbound.persistence.sqlalchemy.oferta_repository import (
    SqlAlchemyOfertaRepository,
)
from src.assistente.models.assistente_models import (
    AssClienteOfertaModel, AssOfertaMensalModel,
    AssRendaFixaModel, AssPosicaoFundoModel, AssFundoCVMModel,
)
from src.assistente.domain.entities.lead import EstagioLead, Lead, OrigemLead
from src.assistente.domain.entities.anotacao import Anotacao, TipoAnotacao

logger = structlog.get_logger(__name__)
router = APIRouter(prefix="/assistente", tags=["Assistente"])


# ── Schemas ────────────────────────────────────────────────────────────────────

class EventoRequest(BaseModel):
    tipo: str
    descricao: str
    data_evento: date
    alertar_dias_antes: int = 1
    codigo_conta: str | None = None
    nome_cliente: str | None = None


class ReuniaoRequest(BaseModel):
    titulo: str
    data_hora: datetime
    duracao_minutos: int = 60
    codigo_conta: str | None = None
    nome_cliente: str | None = None
    descricao: str | None = None
    gerar_relatorio: bool = False
    criar_no_outlook: bool = True


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/clientes/importar")
async def importar_clientes_endpoint(
    request: Request,
    arquivo_saldo: UploadFile = File(...),
    arquivo_positivador: UploadFile = File(...),
):
    """Importa clientes a partir das planilhas RelatorioSaldoConsolidado e Positivador."""
    import tempfile, shutil

    settings = request.app.state.settings
    session_factory = request.app.state.session_factory

    # Salvar uploads em arquivos temporários
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f_saldo:
        shutil.copyfileobj(arquivo_saldo.file, f_saldo)
        caminho_saldo = f_saldo.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f_pos:
        shutil.copyfileobj(arquivo_positivador.file, f_pos)
        caminho_positivador = f_pos.name

    try:
        clientes = importar_clientes(
            caminho_relatorio_saldo=caminho_saldo,
            caminho_positivador=caminho_positivador,
            codigo_assessor="69567",
        )
        async with session_factory() as session:
            async with session.begin():
                repo = SqlAlchemyClienteAssessorRepository(session)
                total = await repo.salvar_todos(clientes)

        return {"importados": total, "mensagem": f"{total} clientes importados/atualizados"}
    finally:
        Path(caminho_saldo).unlink(missing_ok=True)
        Path(caminho_positivador).unlink(missing_ok=True)


@router.get("/clientes")
async def listar_clientes(request: Request):
    """Lista todos os clientes do assessor."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyClienteAssessorRepository(session)
        clientes = await repo.listar_todos()

    return [
        {
            "codigo_conta": c.codigo_conta,
            "nome": c.nome,
            "net": c.net,
            "suitability": c.suitability,
            "profissao": c.profissao,
            "data_nascimento": str(c.data_nascimento) if c.data_nascimento else None,
            "tem_liquidacao": c.tem_liquidacao_pendente,
        }
        for c in clientes
    ]


@router.post("/eventos")
async def registrar_evento(request: Request, body: EventoRequest):
    """Registra uma data importante para monitoramento."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyEventoRepository(session)
            handler = RegistrarEventoHandler(evento_repository=repo)
            evento_id = await handler.handle(
                RegistrarEventoCommand(
                    tipo=body.tipo,
                    descricao=body.descricao,
                    data_evento=body.data_evento,
                    alertar_dias_antes=body.alertar_dias_antes,
                    codigo_conta=body.codigo_conta,
                    nome_cliente=body.nome_cliente,
                )
            )
    return {"id": evento_id, "mensagem": "Evento registrado com sucesso"}


@router.get("/eventos/proximos")
async def listar_proximos_eventos(request: Request, dias: int = 7):
    """Lista eventos dos próximos N dias."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyEventoRepository(session)
        eventos = await repo.listar_proximos(dias=dias)

    return [
        {
            "id": e.id,
            "tipo": e.tipo.value,
            "descricao": e.descricao,
            "data": str(e.data_evento),
            "dias_para_evento": e.dias_para_evento(),
            "cliente": e.nome_cliente,
            "conta": e.codigo_conta,
        }
        for e in eventos
    ]


@router.patch("/eventos/{evento_id}/concluir")
async def concluir_evento(request: Request, evento_id: str):
    """Marca um evento/tarefa como concluído."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyEventoRepository(session).marcar_concluido(evento_id)
    return {"mensagem": "Evento concluído"}


@router.delete("/eventos/{evento_id}")
async def deletar_evento(request: Request, evento_id: str):
    """Remove um evento/tarefa."""
    from sqlalchemy import delete as sa_delete
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            from src.assistente.models.assistente_models import AssEventoModel
            await session.execute(sa_delete(AssEventoModel).where(AssEventoModel.id == evento_id))
    return {"mensagem": "Evento removido"}


@router.post("/reunioes")
async def agendar_reuniao(request: Request, body: ReuniaoRequest):
    """Agenda reunião com cliente (cria no Outlook se configurado)."""
    settings = request.app.state.settings
    session_factory = request.app.state.session_factory

    calendar_port = None
    if settings.ms_tenant_id and body.criar_no_outlook:
        from src.assistente.adapters.outbound.calendar.outlook_calendar_adapter import (
            OutlookCalendarAdapter,
        )
        calendar_port = OutlookCalendarAdapter(
            tenant_id=settings.ms_tenant_id,
            client_id=settings.ms_client_id,
            client_secret=settings.ms_client_secret,
            user_email=settings.ms_user_email,
        )

    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyReuniaoRepository(session)
            handler = AgendarReuniaoHandler(
                reuniao_repository=repo,
                calendar_port=calendar_port,
            )
            reuniao_id = await handler.handle(
                AgendarReuniaoCommand(
                    titulo=body.titulo,
                    data_hora=body.data_hora,
                    duracao_minutos=body.duracao_minutos,
                    codigo_conta=body.codigo_conta,
                    nome_cliente=body.nome_cliente,
                    descricao=body.descricao,
                    gerar_relatorio=body.gerar_relatorio,
                    criar_no_outlook=body.criar_no_outlook and calendar_port is not None,
                )
            )
    return {"id": reuniao_id, "mensagem": "Reunião agendada com sucesso"}


@router.get("/reunioes/proximas")
async def listar_proximas_reunioes(request: Request, dias: int = 7):
    """Lista reuniões agendadas nos próximos N dias."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyReuniaoRepository(session)
        reunioes = await repo.listar_proximas(dias=dias)

    return [
        {
            "id": r.id,
            "titulo": r.titulo,
            "data_hora": r.data_hora.isoformat(),
            "duracao_minutos": r.duracao_minutos,
            "descricao": r.descricao,
            "cliente": r.nome_cliente,
            "conta": r.codigo_conta,
            "gerar_relatorio": r.gerar_relatorio,
        }
        for r in reunioes
    ]


@router.patch("/reunioes/{reuniao_id}/cancelar")
async def cancelar_reuniao(request: Request, reuniao_id: str):
    """Cancela uma reunião agendada."""
    from sqlalchemy import update as sa_update
    from src.assistente.models.assistente_models import AssReuniaoModel
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await session.execute(
                sa_update(AssReuniaoModel)
                .where(AssReuniaoModel.id == reuniao_id)
                .values(status="CANCELADA")
            )
    return {"mensagem": "Reunião cancelada"}


# ── Diversificador / Renda Fixa ───────────────────────────────────────────────

@router.post("/diversificador/importar")
async def importar_diversificador(
    request: Request,
    arquivo: UploadFile = File(...),
    alertar_dias_antes: int = 30,
):
    """Importa o Diversificador XP e registra posições de renda fixa com alertas de vencimento.

    Aceita o arquivo Diversificacao-A{assessor}-Ref.{data}.xlsx.
    Substitui todas as posições anteriores e recria os eventos de vencimento.
    """
    import tempfile, shutil

    session_factory = request.app.state.session_factory

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        shutil.copyfileobj(arquivo.file, f)
        caminho_temp = f.name

    try:
        async with session_factory() as session:
            async with session.begin():
                rf_repo = SqlAlchemyRendaFixaRepository(session)
                evento_repo = SqlAlchemyEventoRepository(session)
                cliente_repo = SqlAlchemyClienteAssessorRepository(session)
                historico_repo = SqlAlchemyHistoricoImportsRepository(session)
                handler = ImportarDiversificadorHandler(
                    renda_fixa_repository=rf_repo,
                    evento_repository=evento_repo,
                    cliente_assessor_repository=cliente_repo,
                    historico_repo=historico_repo,
                )
                resultado = await handler.handle(
                    ImportarDiversificadorCommand(
                        caminho_diversificador=caminho_temp,
                        alertar_dias_antes=alertar_dias_antes,
                        substituir_existentes=True,
                    )
                )
    finally:
        Path(caminho_temp).unlink(missing_ok=True)

    return {
        "posicoes_importadas": resultado.posicoes_importadas,
        "eventos_criados": resultado.eventos_criados,
        "clientes_com_rf": resultado.clientes_com_rf,
        "data_referencia": str(resultado.data_referencia) if resultado.data_referencia else None,
        "mensagem": (
            f"{resultado.posicoes_importadas} posições RF importadas, "
            f"{resultado.eventos_criados} alertas de vencimento criados."
        ),
    }


@router.get("/diversificador/vencimentos")
async def listar_vencimentos(request: Request, dias: int = 90):
    """Lista posições de renda fixa vencendo nos próximos N dias."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyRendaFixaRepository(session)
        posicoes = await repo.listar_vencendo_em(dias=dias)

    return [
        {
            "conta": p.codigo_conta,
            "cliente": p.nome_cliente,
            "tipo": p.tipo_ativo,
            "ativo": p.dsc_ativo,
            "emissor": p.emissor,
            "indexador": p.indexador,
            "vencimento": str(p.data_vencimento),
            "valor": p.valor_aplicado,
        }
        for p in posicoes
    ]


@router.get("/diversificador/posicoes/{codigo_conta}")
async def listar_posicoes_cliente(request: Request, codigo_conta: str):
    """Lista todas as posições de renda fixa de um cliente."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyRendaFixaRepository(session)
        posicoes = await repo.listar_por_conta(codigo_conta)

    return [
        {
            "tipo": p.tipo_ativo,
            "ativo": p.dsc_ativo,
            "emissor": p.emissor,
            "indexador": p.indexador,
            "vencimento": str(p.data_vencimento),
            "valor": p.valor_aplicado,
        }
        for p in posicoes
    ]


# ── Diversificador Completo (RV + Fundos + Previdência) ───────────────────────

@router.post("/diversificador/importar-completo")
async def importar_diversificador_completo(
    request: Request,
    arquivo: UploadFile = File(...),
):
    """Importa o Diversificador XP: acoes, FIIs, fundos e previdencia."""
    import tempfile, shutil
    session_factory = request.app.state.session_factory

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        shutil.copyfileobj(arquivo.file, f)
        caminho_temp = f.name

    try:
        async with session_factory() as session:
            async with session.begin():
                handler = ImportarDiversificadorCompletoHandler(
                    rv_repo=SqlAlchemyPosicaoRVRepository(session),
                    fundo_repo=SqlAlchemyPosicaoFundoRepository(session),
                    prev_repo=SqlAlchemyPosicaoPrevRepository(session),
                    cliente_repo=SqlAlchemyClienteAssessorRepository(session),
                    historico_repo=SqlAlchemyHistoricoImportsRepository(session),
                )
                resultado = await handler.handle(caminho_temp)
    finally:
        Path(caminho_temp).unlink(missing_ok=True)

    return {
        "rv": resultado.rv,
        "fundos": resultado.fundos,
        "prev": resultado.prev,
        "data_referencia": resultado.data_referencia,
        "mensagem": f"{resultado.rv} posicoes RV, {resultado.fundos} fundos, {resultado.prev} previdencia importados.",
    }


@router.get("/diversificador/posicoes-rv/{codigo_conta}")
async def listar_posicoes_rv_cliente(request: Request, codigo_conta: str):
    """Lista posicoes de renda variavel (acoes, FIIs, opcoes) de um cliente."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        posicoes = await SqlAlchemyPosicaoRVRepository(session).listar_por_conta(codigo_conta)
    return [
        {"tipo": p.tipo, "ticker": p.ticker, "ativo": p.dsc_ativo,
         "quantidade": p.quantidade, "valor": p.valor_net}
        for p in posicoes
    ]


@router.get("/diversificador/fundos/{codigo_conta}")
async def listar_fundos_cliente(request: Request, codigo_conta: str):
    """Lista fundos de investimento de um cliente (com prazo de resgate CVM)."""
    from sqlalchemy import select
    from src.assistente.models.assistente_models import AssFundoCVMModel
    from src.assistente.adapters.outbound.services.cvm_fundo_service import (
        _normalizar_cnpj, formatar_prazo,
    )
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        posicoes = await SqlAlchemyPosicaoFundoRepository(session).listar_por_conta(codigo_conta)
        cnpjs = {_normalizar_cnpj(p.cnpj_fundo) for p in posicoes if p.cnpj_fundo}
        cvm_map: dict[str, any] = {}
        if cnpjs:
            rows = (await session.execute(
                select(AssFundoCVMModel).where(AssFundoCVMModel.cnpj.in_(cnpjs))
            )).scalars().all()
            cvm_map = {r.cnpj: r for r in rows}
    result = []
    for p in posicoes:
        cnpj_norm = _normalizar_cnpj(p.cnpj_fundo)
        cvm = cvm_map.get(cnpj_norm)
        result.append({
            "tipo": p.tipo_fundo,
            "fundo": p.nome_fundo,
            "gestora": p.gestora,
            "valor": p.valor_net,
            "cnpj": p.cnpj_fundo,
            "prazo_cotiz": cvm.prazo_cotiz_resg if cvm else None,
            "prazo_pagto": cvm.prazo_pagto_resg if cvm else None,
            "tipo_dia_cotiz": cvm.tipo_dia_cotiz if cvm else None,
            "tipo_dia_pagto": cvm.tipo_dia_pagto if cvm else None,
            "prazo_fmt": formatar_prazo(
                cvm.prazo_cotiz_resg, cvm.prazo_pagto_resg,
                cvm.tipo_dia_cotiz, cvm.tipo_dia_pagto,
            ) if cvm else "—",
        })
    return result


@router.get("/diversificador/previdencia/{codigo_conta}")
async def listar_prev_cliente(request: Request, codigo_conta: str):
    """Lista posicoes de previdencia privada de um cliente (com prazo de resgate CVM)."""
    from sqlalchemy import select
    from src.assistente.models.assistente_models import AssFundoCVMModel
    from src.assistente.adapters.outbound.services.cvm_fundo_service import (
        _normalizar_cnpj, formatar_prazo,
    )
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        posicoes = await SqlAlchemyPosicaoPrevRepository(session).listar_por_conta(codigo_conta)
        # Previdência pode não ter CNPJ; tenta pelo nome se tiver mapeamento futuro
        cnpjs = {_normalizar_cnpj(getattr(p, "cnpj_fundo", None)) for p in posicoes if getattr(p, "cnpj_fundo", None)}
        cvm_map: dict[str, any] = {}
        if cnpjs:
            rows = (await session.execute(
                select(AssFundoCVMModel).where(AssFundoCVMModel.cnpj.in_(cnpjs))
            )).scalars().all()
            cvm_map = {r.cnpj: r for r in rows}
    result = []
    for p in posicoes:
        cnpj_norm = _normalizar_cnpj(getattr(p, "cnpj_fundo", None))
        cvm = cvm_map.get(cnpj_norm)
        result.append({
            "tipo": p.tipo_fundo,
            "fundo": p.nome_fundo,
            "gestora": p.gestora,
            "valor": p.valor_net,
            "prazo_fmt": formatar_prazo(
                cvm.prazo_cotiz_resg, cvm.prazo_pagto_resg,
                cvm.tipo_dia_cotiz, cvm.tipo_dia_pagto,
            ) if cvm else "—",
        })
    return result


# ── Fundos: Import Lista XP ───────────────────────────────────────────────────

@router.post("/fundos/importar-lista-xp")
async def importar_lista_xp(request: Request, arquivo: UploadFile = File(...)):
    """Importa a planilha lista-fundos-*.xlsx da XP com prazos de resgate."""
    import tempfile, shutil
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from src.assistente.models.assistente_models import AssFundoCVMModel
    from src.assistente.adapters.outbound.services.cvm_fundo_service import importar_lista_xp

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        shutil.copyfileobj(arquivo.file, f)
        caminho_temp = f.name

    try:
        fundos = importar_lista_xp(caminho_temp)
    except (ValueError, ImportError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        Path(caminho_temp).unlink(missing_ok=True)

    if not fundos:
        raise HTTPException(status_code=400, detail="Nenhum fundo encontrado na planilha.")

    agora = datetime.now(timezone.utc)
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            for f in fundos:
                stmt = pg_insert(AssFundoCVMModel).values(
                    cnpj=f["cnpj"],
                    denom_social=f["denom_social"],
                    gestora=f["gestora"],
                    situacao=f["situacao"],
                    prazo_cotiz_resg=f["prazo_cotiz_resg"],
                    tipo_dia_cotiz=f["tipo_dia_cotiz"],
                    prazo_pagto_resg=f["prazo_pagto_resg"],
                    tipo_dia_pagto=f["tipo_dia_pagto"],
                    atualizado_em=agora,
                ).on_conflict_do_update(
                    index_elements=["cnpj"],
                    set_={k: f[k] for k in (
                        "denom_social", "gestora", "situacao",
                        "prazo_cotiz_resg", "tipo_dia_cotiz",
                        "prazo_pagto_resg", "tipo_dia_pagto", "atualizado_em",
                    )},
                )
                await session.execute(stmt)

    return {"importados": len(fundos), "mensagem": f"{len(fundos)} fundos importados com sucesso."}


# ── CVM: Prazos de Resgate ─────────────────────────────────────────────────────

@router.post("/fundos/atualizar-prazos-cvm")
async def atualizar_prazos_cvm(request: Request, forcar: bool = False):
    """Busca prazos de resgate na CVM apenas para fundos ainda não cadastrados.

    Usa o cache em ass_fundos_cvm — só chama a CVM se houver CNPJs novos
    (ou se forcar=true para atualizar tudo).
    """
    from sqlalchemy import select
    from src.assistente.models.assistente_models import AssFundoCVMModel, AssPosicaoFundoModel
    from src.assistente.adapters.outbound.services.cvm_fundo_service import (
        buscar_prazos_cvm, _normalizar_cnpj,
    )
    session_factory = request.app.state.session_factory

    async with session_factory() as session:
        # CNPJs que temos em posições de fundos
        fundos_rows = (await session.execute(
            select(AssPosicaoFundoModel.cnpj_fundo).distinct()
        )).scalars().all()

        # CNPJs já salvos na tabela CVM
        ja_salvos = set(
            (await session.execute(select(AssFundoCVMModel.cnpj))).scalars().all()
        )

    cnpjs_norm = {_normalizar_cnpj(c) for c in fundos_rows if _normalizar_cnpj(c)}

    if not cnpjs_norm:
        return {"mensagem": "Nenhum fundo com CNPJ encontrado na base.", "novos": 0, "ja_na_cache": 0}

    # Se não forçar, filtra só os que ainda não estão no cache
    cnpjs_novos = cnpjs_norm if forcar else (cnpjs_norm - ja_salvos)

    ja_na_cache = len(cnpjs_norm) - len(cnpjs_novos)

    if not cnpjs_novos:
        return {
            "mensagem": f"Todos os {ja_na_cache} fundos já estão em cache. Use forcar=true para atualizar.",
            "novos": 0,
            "ja_na_cache": ja_na_cache,
        }

    # Buscar na CVM apenas os CNPJs novos
    try:
        dados_cvm = await buscar_prazos_cvm(cnpjs_novos)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))

    from sqlalchemy.dialects.postgresql import insert as pg_insert
    agora = datetime.now(timezone.utc)
    async with session_factory() as session:
        async with session.begin():
            for cnpj, info in dados_cvm.items():
                stmt = pg_insert(AssFundoCVMModel).values(
                    cnpj=cnpj,
                    denom_social=info["denom_social"],
                    situacao=info["situacao"],
                    prazo_cotiz_resg=info["prazo_cotiz_resg"],
                    prazo_pagto_resg=info["prazo_pagto_resg"],
                    atualizado_em=agora,
                ).on_conflict_do_update(
                    index_elements=["cnpj"],
                    set_={
                        "denom_social": info["denom_social"],
                        "situacao": info["situacao"],
                        "prazo_cotiz_resg": info["prazo_cotiz_resg"],
                        "prazo_pagto_resg": info["prazo_pagto_resg"],
                        "atualizado_em": agora,
                    },
                )
                await session.execute(stmt)

    return {
        "novos": len(dados_cvm),
        "ja_na_cache": ja_na_cache,
        "mensagem": (
            f"{len(dados_cvm)} novo(s) fundo(s) buscado(s) na CVM. "
            f"{ja_na_cache} já estavam em cache."
        ),
    }


@router.get("/fundos")
async def listar_todos_fundos(request: Request):
    """Lista todas as posições de fundos de todos os clientes com prazo de resgate."""
    from sqlalchemy import select
    from src.assistente.models.assistente_models import (
        AssPosicaoFundoModel, AssPosicaoPrevModel, AssFundoCVMModel,
    )
    from src.assistente.adapters.outbound.services.cvm_fundo_service import (
        _normalizar_cnpj, formatar_prazo,
    )
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        fundos = (await session.execute(
            select(AssPosicaoFundoModel).order_by(AssPosicaoFundoModel.valor_net.desc())
        )).scalars().all()
        prev = (await session.execute(
            select(AssPosicaoPrevModel).order_by(AssPosicaoPrevModel.valor_net.desc())
        )).scalars().all()

        # Carregar todos os registros CVM de uma vez
        all_cvm = (await session.execute(select(AssFundoCVMModel))).scalars().all()
        cvm_map = {r.cnpj: r for r in all_cvm}

    def _enrich(p: any, classe: str) -> dict:
        cnpj_norm = _normalizar_cnpj(getattr(p, "cnpj_fundo", None))
        cvm = cvm_map.get(cnpj_norm)
        return {
            "classe": classe,
            "codigo_conta": p.codigo_conta,
            "nome_cliente": p.nome_cliente,
            "tipo": p.tipo_fundo,
            "fundo": p.nome_fundo,
            "gestora": p.gestora,
            "valor": p.valor_net,
            "cnpj": getattr(p, "cnpj_fundo", None),
            "prazo_cotiz": cvm.prazo_cotiz_resg if cvm else None,
            "prazo_pagto": cvm.prazo_pagto_resg if cvm else None,
            "tipo_dia_cotiz": cvm.tipo_dia_cotiz if cvm else None,
            "tipo_dia_pagto": cvm.tipo_dia_pagto if cvm else None,
            "prazo_fmt": formatar_prazo(
                cvm.prazo_cotiz_resg, cvm.prazo_pagto_resg,
                cvm.tipo_dia_cotiz, cvm.tipo_dia_pagto,
            ) if cvm else "—",
        }

    resultado = [_enrich(p, "FUNDO") for p in fundos] + [_enrich(p, "PREV") for p in prev]
    resultado.sort(key=lambda x: x["valor"] or 0, reverse=True)
    return resultado


@router.get("/diversificador/buscar-por-ticker")
async def buscar_clientes_por_ticker(request: Request, ticker: str):
    """Retorna todos os clientes que possuem determinado ativo na carteira de RV."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        posicoes = await SqlAlchemyPosicaoRVRepository(session).listar_clientes_por_ticker(ticker)
    return [
        {
            "codigo_conta": p.codigo_conta,
            "nome_cliente": p.nome_cliente,
            "ticker": p.ticker,
            "tipo": p.tipo,
            "quantidade": p.quantidade,
            "valor": p.valor_net,
        }
        for p in posicoes
    ]


# ── Dashboard: Exposição Consolidada ──────────────────────────────────────────

@router.get("/diversificador/resumo-produto")
async def resumo_por_produto(request: Request):
    """Totais por classe de ativo para o painel Resumo por Produto."""
    from sqlalchemy import select, func
    from src.assistente.models.assistente_models import (
        AssRendaFixaModel, AssPosicaoRVModel, AssPosicaoFundoModel, AssPosicaoPrevModel,
    )
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        rf_total = (await session.execute(
            select(func.sum(AssRendaFixaModel.valor_aplicado))
        )).scalar() or 0.0

        rv_rows = (await session.execute(
            select(AssPosicaoRVModel.tipo, func.sum(AssPosicaoRVModel.valor_net).label("total"))
            .group_by(AssPosicaoRVModel.tipo)
        )).all()
        rv_by_tipo = {r.tipo: (r.total or 0.0) for r in rv_rows}

        fundos_rows = (await session.execute(
            select(AssPosicaoFundoModel.tipo_fundo, func.sum(AssPosicaoFundoModel.valor_net).label("total"))
            .group_by(AssPosicaoFundoModel.tipo_fundo)
        )).all()
        fundos_by_tipo = {r.tipo_fundo: (r.total or 0.0) for r in fundos_rows}

        prev_total = (await session.execute(
            select(func.sum(AssPosicaoPrevModel.valor_net))
        )).scalar() or 0.0

    resultado = [
        {"classe": "Renda Fixa",      "valor": rf_total},
        {"classe": "Ações",           "valor": rv_by_tipo.get("ACAO", 0.0)},
        {"classe": "FII",             "valor": rv_by_tipo.get("FII", 0.0)},
        {"classe": "ETF",             "valor": rv_by_tipo.get("ETF", 0.0)},
        {"classe": "Fundo RF",        "valor": fundos_by_tipo.get("RF", 0.0)},
        {"classe": "Fundo Multim.",   "valor": fundos_by_tipo.get("MULTIMERCADO", 0.0)},
        {"classe": "Fundo Ações",     "valor": fundos_by_tipo.get("ACOES", 0.0)},
        {"classe": "Previdência",     "valor": prev_total},
    ]
    total = sum(r["valor"] for r in resultado)
    for r in resultado:
        r["pct"] = round(r["valor"] / total * 100, 1) if total > 0 else 0.0
    return [r for r in resultado if r["valor"] > 0]


@router.get("/diversificador/exposicao-rf-fundos")
async def exposicao_rf_fundos(request: Request):
    """Top 20 maiores posições somadas de RF + Fundos + Previdência."""
    from sqlalchemy import select, func
    from src.assistente.models.assistente_models import (
        AssRendaFixaModel, AssPosicaoFundoModel, AssPosicaoPrevModel,
    )
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        rf_rows = (await session.execute(
            select(
                AssRendaFixaModel.tipo_ativo.label("categoria"),
                AssRendaFixaModel.emissor.label("nome"),
                func.sum(AssRendaFixaModel.valor_aplicado).label("total"),
            ).group_by(AssRendaFixaModel.tipo_ativo, AssRendaFixaModel.emissor)
        )).all()
        rf = [{"classe": "RF", "categoria": r.categoria, "nome": r.nome or "—", "total": r.total or 0} for r in rf_rows]

        f_rows = (await session.execute(
            select(
                AssPosicaoFundoModel.tipo_fundo.label("categoria"),
                AssPosicaoFundoModel.nome_fundo.label("nome"),
                func.sum(AssPosicaoFundoModel.valor_net).label("total"),
            ).group_by(AssPosicaoFundoModel.tipo_fundo, AssPosicaoFundoModel.nome_fundo)
        )).all()
        fundos = [{"classe": "FUNDO", "categoria": r.categoria, "nome": r.nome or "—", "total": r.total or 0} for r in f_rows]

        p_rows = (await session.execute(
            select(
                AssPosicaoPrevModel.tipo_fundo.label("categoria"),
                AssPosicaoPrevModel.nome_fundo.label("nome"),
                func.sum(AssPosicaoPrevModel.valor_net).label("total"),
            ).group_by(AssPosicaoPrevModel.tipo_fundo, AssPosicaoPrevModel.nome_fundo)
        )).all()
        prev = [{"classe": "PREV", "categoria": r.categoria, "nome": r.nome or "—", "total": r.total or 0} for r in p_rows]

    combinado = sorted(rf + fundos + prev, key=lambda x: x["total"], reverse=True)[:20]
    return combinado


@router.get("/diversificador/exposicao-rv")
async def exposicao_rv(request: Request):
    """Top 20 maiores posições somadas de Ações + FII + ETF por ticker."""
    from sqlalchemy import select, func
    from src.assistente.models.assistente_models import AssPosicaoRVModel
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        rows = (await session.execute(
            select(
                AssPosicaoRVModel.ticker,
                AssPosicaoRVModel.tipo,
                AssPosicaoRVModel.dsc_ativo.label("nome"),
                func.sum(AssPosicaoRVModel.valor_net).label("total"),
            )
            .where(AssPosicaoRVModel.tipo.in_(["ACAO", "FII", "ETF", "BDR"]))
            .group_by(AssPosicaoRVModel.ticker, AssPosicaoRVModel.tipo, AssPosicaoRVModel.dsc_ativo)
            .order_by(func.sum(AssPosicaoRVModel.valor_net).desc())
            .limit(20)
        )).all()
    return [{"ticker": r.ticker or "—", "tipo": r.tipo, "nome": r.nome or "—", "total": r.total or 0} for r in rows]


# ── Leads / Funil de Prospeccao ───────────────────────────────────────────────

class LeadRequest(BaseModel):
    nome: str
    telefone: str | None = None
    email: str | None = None
    origem: str | None = None
    valor_potencial: float | None = None
    anotacoes: str | None = None
    proximo_passo: str | None = None
    data_proximo_contato: date | None = None


@router.get("/leads")
async def listar_leads(request: Request, estagio: str | None = None):
    """Lista leads. Filtrar por estagio: PROSPECTO | CONTATO | PROPOSTA | CLIENTE."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyLeadRepository(session)
        leads = await repo.listar_por_estagio(estagio) if estagio else await repo.listar_todos()
    return [
        {"id": l.id, "nome": l.nome, "estagio": l.estagio.value,
         "telefone": l.telefone, "email": l.email,
         "origem": l.origem.value if l.origem else None,
         "valor_potencial": l.valor_potencial, "proximo_passo": l.proximo_passo,
         "data_proximo_contato": str(l.data_proximo_contato) if l.data_proximo_contato else None}
        for l in leads
    ]


@router.post("/leads")
async def criar_lead(request: Request, body: LeadRequest):
    """Cria um novo lead no funil."""
    import uuid as _uuid
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyLeadRepository(session)
            lead = Lead(
                id=str(_uuid.uuid4()),
                nome=body.nome,
                telefone=body.telefone,
                email=body.email,
                origem=OrigemLead(body.origem) if body.origem else None,
                valor_potencial=body.valor_potencial,
                anotacoes=body.anotacoes,
                proximo_passo=body.proximo_passo,
                data_proximo_contato=body.data_proximo_contato,
            )
            await repo.salvar(lead)
    return {"id": lead.id, "mensagem": "Lead criado"}


@router.patch("/leads/{lead_id}/estagio")
async def atualizar_estagio_lead(request: Request, lead_id: str, estagio: str):
    """Move o lead para outro estagio do funil."""
    estagios_validos = [e.value for e in EstagioLead]
    if estagio not in estagios_validos:
        raise HTTPException(400, f"Estagio invalido. Use: {estagios_validos}")
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyLeadRepository(session).atualizar_estagio(lead_id, estagio)
    return {"mensagem": f"Lead movido para {estagio}"}


@router.patch("/leads/{lead_id}")
async def atualizar_lead(request: Request, lead_id: str, body: LeadRequest):
    """Atualiza dados de um lead."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyLeadRepository(session)
            lead = await repo.buscar_por_id(lead_id)
            if not lead:
                raise HTTPException(404, "Lead nao encontrado")
            lead.nome = body.nome
            lead.telefone = body.telefone
            lead.email = body.email
            lead.origem = OrigemLead(body.origem) if body.origem else None
            lead.valor_potencial = body.valor_potencial
            lead.anotacoes = body.anotacoes
            lead.proximo_passo = body.proximo_passo
            lead.data_proximo_contato = body.data_proximo_contato
            await repo.salvar(lead)
    return {"mensagem": "Lead atualizado"}


@router.delete("/leads/{lead_id}")
async def deletar_lead(request: Request, lead_id: str):
    """Remove um lead."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyLeadRepository(session).deletar(lead_id)
    return {"mensagem": "Lead removido"}


@router.get("/leads/template-excel")
async def download_template_leads():
    """Gera e retorna o template Excel para importação de leads."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise HTTPException(500, "openpyxl não instalado") from e

    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["nome", "telefone", "email", "origem", "valor_potencial", "estagio", "anotacoes"]
    header_labels = ["Nome *", "Telefone", "Email", "Origem", "Patrimônio Potencial (R$)", "Estágio", "Anotações"]

    # Cabeçalho estilizado
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # Linha de exemplo
    ws.append(["João Silva", "11999990000", "joao@email.com", "INDICACAO", 500000, "PROSPECTO", "Indicado pelo cliente X"])

    # Aba de referência com valores válidos
    ws_ref = wb.create_sheet("Referência")
    ws_ref["A1"] = "Origem (valores válidos)"
    ws_ref["B1"] = "Estágio (valores válidos)"
    for i, v in enumerate(["INDICACAO", "EVENTO", "LINKEDIN", "COLD_CALL", "OUTRO"], 2):
        ws_ref[f"A{i}"] = v
    for i, v in enumerate(["PROSPECTO", "CONTATO", "PROPOSTA", "CLIENTE"], 2):
        ws_ref[f"B{i}"] = v

    # Ajustar larguras
    for col, w in zip("ABCDEFG", [30, 15, 25, 15, 25, 12, 40]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_leads.xlsx"},
    )


@router.post("/leads/importar-excel")
async def importar_leads_excel(request: Request, arquivo: UploadFile = File(...)):
    """Importa leads a partir de planilha Excel (usando o template disponível para download)."""
    try:
        import openpyxl
    except ImportError as e:
        raise HTTPException(500, "openpyxl não instalado") from e

    import io, uuid as _uuid
    from src.assistente.domain.entities.lead import Lead, EstagioLead, OrigemLead

    ORIGENS_VALIDAS = {o.value for o in OrigemLead}
    ESTAGIOS_VALIDOS = {e.value for e in EstagioLead}

    conteudo = await arquivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    leads: list[Lead] = []
    erros: list[str] = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        nome        = str(row[0]).strip() if row[0] else None
        telefone    = str(row[1]).strip() if row[1] else None
        email       = str(row[2]).strip() if row[2] else None
        origem_raw  = str(row[3]).strip().upper() if row[3] else None
        val_pot     = float(row[4]) if row[4] else None
        estagio_raw = str(row[5]).strip().upper() if row[5] else "PROSPECTO"
        anotacoes   = str(row[6]).strip() if row[6] else None

        if not nome:
            erros.append(f"Linha {i}: nome obrigatório")
            continue

        origem  = OrigemLead(origem_raw)  if origem_raw  in ORIGENS_VALIDAS  else None
        estagio = EstagioLead(estagio_raw) if estagio_raw in ESTAGIOS_VALIDOS else EstagioLead.PROSPECTO

        leads.append(Lead(
            id=str(_uuid.uuid4()),
            nome=nome,
            telefone=telefone,
            email=email,
            origem=origem,
            valor_potencial=val_pot,
            estagio=estagio,
            anotacoes=anotacoes,
        ))

    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyLeadRepository(session)
            for lead in leads:
                await repo.salvar(lead)

    return {
        "importados": len(leads),
        "erros": erros,
        "mensagem": f"{len(leads)} leads importados" + (f" ({len(erros)} linhas ignoradas)" if erros else ""),
    }


# ── Anotacoes / Timeline de contatos ─────────────────────────────────────────

class AnotacaoRequest(BaseModel):
    tipo: str
    texto: str


@router.get("/clientes/{codigo_conta}/anotacoes")
async def listar_anotacoes(request: Request, codigo_conta: str):
    """Retorna timeline de contatos de um cliente."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        anotacoes = await SqlAlchemyAnotacaoRepository(session).listar_por_conta(codigo_conta)
    return [
        {"id": a.id, "tipo": a.tipo.value, "texto": a.texto, "criado_em": a.criado_em.isoformat()}
        for a in anotacoes
    ]


@router.post("/clientes/{codigo_conta}/anotacoes")
async def criar_anotacao(request: Request, codigo_conta: str, body: AnotacaoRequest):
    """Registra um contato ou nota na timeline do cliente."""
    import uuid as _uuid
    tipos_validos = [t.value for t in TipoAnotacao]
    if body.tipo not in tipos_validos:
        raise HTTPException(400, f"Tipo invalido. Use: {tipos_validos}")
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyAnotacaoRepository(session)
            anotacao = Anotacao(
                id=str(_uuid.uuid4()),
                codigo_conta=codigo_conta,
                tipo=TipoAnotacao(body.tipo),
                texto=body.texto,
            )
            await repo.salvar(anotacao)
    return {"id": anotacao.id, "mensagem": "Anotacao registrada"}


@router.delete("/anotacoes/{anotacao_id}")
async def deletar_anotacao(request: Request, anotacao_id: str):
    """Remove uma anotacao."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyAnotacaoRepository(session).deletar(anotacao_id)
    return {"mensagem": "Anotacao removida"}


# ── Histórico de Imports ───────────────────────────────────────────────────────

@router.post("/clientes/gerar-alertas-aniversario")
async def gerar_alertas_aniversario(request: Request):
    """Gera (ou atualiza) eventos de aniversário para os próximos 30 dias."""
    from src.assistente.adapters.inbound.scheduler.daily_jobs import AssistenteDailyJobs
    session_factory = request.app.state.session_factory
    settings = request.app.state.settings
    jobs = AssistenteDailyJobs(session_factory=session_factory, settings=settings)
    async with session_factory() as session:
        async with session.begin():
            await jobs._gerar_alertas_aniversario(session)
    return {"mensagem": "Alertas de aniversário gerados com sucesso"}


@router.get("/historico-imports")
async def listar_historico_imports(request: Request):
    """Retorna o histórico dos últimos imports (máximo 3 por tipo)."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        rows = await SqlAlchemyHistoricoImportsRepository(session).listar_todos()
    return [
        {
            "id": r.id,
            "tipo": r.tipo,
            "data_referencia": str(r.data_referencia) if r.data_referencia else None,
            "total_rf": r.total_rf,
            "total_rv": r.total_rv,
            "total_fundos": r.total_fundos,
            "total_prev": r.total_prev,
            "total_clientes": r.total_clientes,
            "importado_em": r.importado_em.isoformat(),
        }
        for r in rows
    ]


# ── Import Unificado (todas as classes num único upload) ──────────────────────

@router.post("/diversificador/importar-tudo")
async def importar_diversificador_tudo(
    request: Request,
    arquivo: UploadFile = File(...),
    alertar_dias_antes: int = 30,
):
    """Importa o Diversificador XP de uma vez: RF + RV + Fundos + Previdência."""
    import tempfile, shutil

    session_factory = request.app.state.session_factory

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        shutil.copyfileobj(arquivo.file, f)
        caminho_temp = f.name

    try:
        async with session_factory() as session:
            async with session.begin():
                historico_repo = SqlAlchemyHistoricoImportsRepository(session)

                # 1. Importar RF
                rf_repo = SqlAlchemyRendaFixaRepository(session)
                evento_repo = SqlAlchemyEventoRepository(session)
                cliente_repo = SqlAlchemyClienteAssessorRepository(session)
                rf_handler = ImportarDiversificadorHandler(
                    renda_fixa_repository=rf_repo,
                    evento_repository=evento_repo,
                    cliente_assessor_repository=cliente_repo,
                )
                rf_resultado = await rf_handler.handle(
                    ImportarDiversificadorCommand(
                        caminho_diversificador=caminho_temp,
                        alertar_dias_antes=alertar_dias_antes,
                        substituir_existentes=True,
                    )
                )

                # 2. Importar RV + Fundos + Prev
                completo_handler = ImportarDiversificadorCompletoHandler(
                    rv_repo=SqlAlchemyPosicaoRVRepository(session),
                    fundo_repo=SqlAlchemyPosicaoFundoRepository(session),
                    prev_repo=SqlAlchemyPosicaoPrevRepository(session),
                    cliente_repo=cliente_repo,
                )
                completo_resultado = await completo_handler.handle(caminho_temp)

                # 3. Registrar histórico único cobrindo tudo
                clientes_distintos = rf_resultado.clientes_com_rf
                await historico_repo.registrar(
                    tipo="TUDO",
                    data_referencia=rf_resultado.data_referencia,
                    total_rf=rf_resultado.posicoes_importadas,
                    total_rv=completo_resultado.rv,
                    total_fundos=completo_resultado.fundos,
                    total_prev=completo_resultado.prev,
                    total_clientes=clientes_distintos,
                )
    finally:
        Path(caminho_temp).unlink(missing_ok=True)

    return {
        "rf": rf_resultado.posicoes_importadas,
        "rv": completo_resultado.rv,
        "fundos": completo_resultado.fundos,
        "prev": completo_resultado.prev,
        "eventos_criados": rf_resultado.eventos_criados,
        "data_referencia": str(rf_resultado.data_referencia) if rf_resultado.data_referencia else completo_resultado.data_referencia,
        "mensagem": (
            f"{rf_resultado.posicoes_importadas} RF · "
            f"{completo_resultado.rv} RV · "
            f"{completo_resultado.fundos} Fundos · "
            f"{completo_resultado.prev} Previdência importados."
        ),
    }


# ── Ofertas Mensais ────────────────────────────────────────────────────────────

_STATUSES_VALIDOS = {"PENDENTE", "WHATSAPP", "RESERVADO", "PUSH_ENVIADO", "FINALIZADO"}


class OfertaRequest(BaseModel):
    nome: str
    descricao: str | None = None
    data_liquidacao: date | None = None
    roa: float = 0.0


class OfertaPatchRequest(BaseModel):
    nome: str | None = None
    descricao: str | None = None
    data_liquidacao: date | None = None
    roa: float | None = None


class ClienteOfertaRequest(BaseModel):
    codigo_conta: str
    nome_cliente: str | None = None
    net: float | None = None
    saldo_disponivel: float | None = None
    valor_ofertado: float | None = None
    status: str = "PENDENTE"
    observacao: str | None = None


@router.get("/ofertas")
async def listar_ofertas(request: Request):
    """Lista todas as ofertas mensais com prévia de receita."""
    import uuid as _uuid
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyOfertaRepository(session)
        ofertas = await repo.listar_ofertas()
        result = []
        for o in ofertas:
            preview = await repo.preview_receita(o.id)
            result.append({
                "id": o.id,
                "nome": o.nome,
                "descricao": o.descricao,
                "data_liquidacao": str(o.data_liquidacao) if o.data_liquidacao else None,
                "roa": o.roa,
                "ativa": o.ativa,
                "criado_em": o.criado_em.isoformat(),
                **preview,
            })
    return result


@router.post("/ofertas")
async def criar_oferta(request: Request, body: OfertaRequest):
    """Cria uma nova oferta mensal."""
    import uuid as _uuid
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyOfertaRepository(session)
            oferta = AssOfertaMensalModel(
                id=str(_uuid.uuid4()),
                nome=body.nome,
                descricao=body.descricao,
                data_liquidacao=body.data_liquidacao,
                roa=body.roa,
            )
            await repo.salvar_oferta(oferta)

            # Criar evento de liquidação se tiver data
            if body.data_liquidacao:
                from src.assistente.domain.entities.evento import Evento, TipoEvento
                evento_repo = SqlAlchemyEventoRepository(session)
                evento = Evento(
                    id=str(_uuid.uuid4()),
                    tipo=TipoEvento.VENCIMENTO_RF,
                    descricao=f"Liquidação oferta: {body.nome}",
                    data_evento=body.data_liquidacao,
                    alertar_dias_antes=5,
                    codigo_conta=None,
                    nome_cliente=None,
                )
                await evento_repo.salvar(evento)

    return {"id": oferta.id, "mensagem": "Oferta criada"}


@router.get("/ofertas/template-excel")
async def download_template_ofertas():
    """Gera e retorna o template Excel para importação de clientes em uma oferta."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise HTTPException(500, "openpyxl não instalado") from e

    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes Oferta"

    header_labels = ["Código Conta *", "Nome do Cliente", "Valor Ofertado (R$)", "Status"]
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    ws.append(["12345678", "João Silva", 50000, "PENDENTE"])

    ws_ref = wb.create_sheet("Referência")
    ws_ref["A1"] = "Status (valores válidos)"
    for i, v in enumerate(["PENDENTE", "WHATSAPP", "RESERVADO", "PUSH_ENVIADO", "FINALIZADO"], 2):
        ws_ref[f"A{i}"] = v

    for col, w in zip("ABCD", [18, 35, 22, 16]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_oferta.xlsx"},
    )


@router.get("/ofertas/{oferta_id}")
async def detalhe_oferta(request: Request, oferta_id: str):
    """Retorna detalhes de uma oferta com todos os clientes e prévia de receita."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyOfertaRepository(session)
        oferta = await repo.buscar_oferta(oferta_id)
        if not oferta:
            raise HTTPException(404, "Oferta não encontrada")
        clientes = await repo.listar_clientes_oferta(oferta_id)
        preview = await repo.preview_receita(oferta_id)
    return {
        "id": oferta.id,
        "nome": oferta.nome,
        "descricao": oferta.descricao,
        "data_liquidacao": str(oferta.data_liquidacao) if oferta.data_liquidacao else None,
        "roa": oferta.roa,
        "ativa": oferta.ativa,
        "criado_em": oferta.criado_em.isoformat(),
        "preview": preview,
        "clientes": [
            {
                "id": c.id,
                "codigo_conta": c.codigo_conta,
                "nome_cliente": c.nome_cliente,
                "net": c.net,
                "saldo_disponivel": c.saldo_disponivel,
                "valor_ofertado": c.valor_ofertado,
                "status": c.status,
                "observacao": c.observacao,
                "atualizado_em": c.atualizado_em.isoformat(),
            }
            for c in clientes
        ],
    }


@router.patch("/ofertas/{oferta_id}")
async def atualizar_oferta(request: Request, oferta_id: str, body: OfertaPatchRequest):
    """Atualiza dados da oferta (campos parciais permitidos)."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyOfertaRepository(session)
            oferta = await repo.buscar_oferta(oferta_id)
            if not oferta:
                raise HTTPException(404, "Oferta não encontrada")
            updates = {k: v for k, v in body.model_dump().items() if v is not None}
            if updates:
                await repo.atualizar_oferta(oferta_id, updates)
    return {"mensagem": "Oferta atualizada"}


@router.delete("/ofertas/{oferta_id}")
async def deletar_oferta(request: Request, oferta_id: str):
    """Remove uma oferta e todos os clientes vinculados."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyOfertaRepository(session).deletar_oferta(oferta_id)
    return {"mensagem": "Oferta removida"}


@router.post("/ofertas/{oferta_id}/clientes")
async def adicionar_cliente_oferta(request: Request, oferta_id: str, body: ClienteOfertaRequest):
    """Adiciona um cliente à oferta."""
    import uuid as _uuid
    if body.status not in _STATUSES_VALIDOS:
        raise HTTPException(400, f"Status inválido. Use: {sorted(_STATUSES_VALIDOS)}")
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyOfertaRepository(session)
            oferta = await repo.buscar_oferta(oferta_id)
            if not oferta:
                raise HTTPException(404, "Oferta não encontrada")
            # Buscar nome do cliente se não fornecido
            nome = body.nome_cliente
            net = body.net
            saldo = body.saldo_disponivel
            if not nome:
                cliente_repo = SqlAlchemyClienteAssessorRepository(session)
                clientes = await cliente_repo.listar_todos()
                cliente = next((c for c in clientes if c.codigo_conta == body.codigo_conta), None)
                if cliente:
                    nome = cliente.nome
                    net = net or cliente.net
            item = AssClienteOfertaModel(
                id=str(_uuid.uuid4()),
                oferta_id=oferta_id,
                codigo_conta=body.codigo_conta,
                nome_cliente=nome,
                net=net,
                saldo_disponivel=saldo,
                valor_ofertado=body.valor_ofertado,
                status=body.status,
                observacao=body.observacao,
            )
            await repo.adicionar_cliente(item)
    return {"id": item.id, "mensagem": "Cliente adicionado à oferta"}


@router.patch("/ofertas/{oferta_id}/clientes/{item_id}")
async def atualizar_cliente_oferta(request: Request, oferta_id: str, item_id: str, body: ClienteOfertaRequest):
    """Atualiza valor ou status de um cliente na oferta."""
    if body.status not in _STATUSES_VALIDOS:
        raise HTTPException(400, f"Status inválido. Use: {sorted(_STATUSES_VALIDOS)}")
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyOfertaRepository(session).atualizar_cliente(item_id, {
                "valor_ofertado": body.valor_ofertado,
                "saldo_disponivel": body.saldo_disponivel,
                "status": body.status,
                "observacao": body.observacao,
            })
    return {"mensagem": "Cliente atualizado"}


@router.delete("/ofertas/{oferta_id}/clientes/{item_id}")
async def remover_cliente_oferta(request: Request, oferta_id: str, item_id: str):
    """Remove um cliente da oferta."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            await SqlAlchemyOfertaRepository(session).remover_cliente(item_id)
    return {"mensagem": "Cliente removido"}


@router.post("/ofertas/{oferta_id}/importar-excel")
async def importar_clientes_oferta_excel(
    request: Request, oferta_id: str, arquivo: UploadFile = File(...)
):
    """Importa clientes para uma oferta a partir de planilha Excel."""
    try:
        import openpyxl
    except ImportError as e:
        raise HTTPException(500, "openpyxl não instalado") from e

    import io
    import uuid as _uuid

    STATUSES_VALIDOS = {"PENDENTE", "WHATSAPP", "RESERVADO", "PUSH_ENVIADO", "FINALIZADO"}

    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Arquivo inválido. Certifique-se de enviar um .xlsx válido. ({e})") from e
    ws = wb.active

    importados = 0
    erros: list[str] = []

    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        async with session.begin():
            repo = SqlAlchemyOfertaRepository(session)

            oferta = await repo.buscar_oferta(oferta_id)
            if not oferta:
                raise HTTPException(404, "Oferta não encontrada")

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                codigo_conta = str(row[0]).strip()
                nome_cliente = str(row[1]).strip() if row[1] else None
                valor_ofertado_raw = row[2]
                status_raw = str(row[3]).strip().upper() if row[3] else "PENDENTE"

                try:
                    valor_ofertado = float(valor_ofertado_raw) if valor_ofertado_raw is not None else None
                except (ValueError, TypeError):
                    erros.append(f"Linha {row_idx}: valor inválido '{valor_ofertado_raw}'")
                    continue

                status = status_raw if status_raw in STATUSES_VALIDOS else "PENDENTE"

                cliente = AssClienteOfertaModel(
                    id=str(_uuid.uuid4()),
                    oferta_id=oferta_id,
                    codigo_conta=codigo_conta,
                    nome_cliente=nome_cliente,
                    valor_ofertado=valor_ofertado,
                    status=status,
                )
                session.add(cliente)
                importados += 1

    return {"importados": importados, "erros": erros}


@router.get("/ofertas/{oferta_id}/preview")
async def preview_oferta(request: Request, oferta_id: str):
    """Retorna prévia de receita da oferta."""
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        return await SqlAlchemyOfertaRepository(session).preview_receita(oferta_id)


@router.get("/ofertas-resumo/clientes")
async def resumo_clientes_ofertas(request: Request):
    """Retorna todos os clientes cruzados com as ofertas, incluindo liquidez calculada."""
    from sqlalchemy import select, func as sa_func
    session_factory = request.app.state.session_factory
    async with session_factory() as session:
        repo = SqlAlchemyOfertaRepository(session)
        cliente_repo = SqlAlchemyClienteAssessorRepository(session)
        todos = await cliente_repo.listar_todos()
        todos_dict = [
            {"codigo_conta": c.codigo_conta, "nome": c.nome, "net": c.net}
            for c in sorted(todos, key=lambda c: c.net or 0, reverse=True)
        ]

        # ── Liquidez: RF com vencimento já passado ─────────────────────────────
        hoje = date.today()
        rf_stmt = (
            select(AssRendaFixaModel.codigo_conta,
                   sa_func.sum(AssRendaFixaModel.valor_aplicado).label("total"))
            .where(AssRendaFixaModel.data_vencimento <= hoje)
            .group_by(AssRendaFixaModel.codigo_conta)
        )
        rf_map: dict[str, float] = {
            r.codigo_conta: (r.total or 0)
            for r in (await session.execute(rf_stmt)).all()
        }

        # ── Liquidez: fundos D+0 e D+1 ────────────────────────────────────────
        fundos_rows = list((await session.execute(select(AssPosicaoFundoModel))).scalars().all())
        cvm_rows = list((await session.execute(select(AssFundoCVMModel))).scalars().all())
        cvm_map = {"".join(c for c in (f.cnpj or "") if c.isdigit()): f for f in cvm_rows}

        fundos_liq_map: dict[str, float] = {}
        fundos_liq_map: dict[str, float] = {}
        fundos_liq6_map: dict[str, float] = {}
        for f in fundos_rows:
            cnpj = "".join(c for c in (f.cnpj_fundo or "") if c.isdigit())
            cvm = cvm_map.get(cnpj)
            if cvm:
                total_prazo = (cvm.prazo_cotiz_resg or 0) + (cvm.prazo_pagto_resg or 0)
                if total_prazo <= 1:
                    fundos_liq_map[f.codigo_conta] = fundos_liq_map.get(f.codigo_conta, 0) + (f.valor_net or 0)
                if total_prazo <= 6:
                    fundos_liq6_map[f.codigo_conta] = fundos_liq6_map.get(f.codigo_conta, 0) + (f.valor_net or 0)

        # ── Combinar RF + Fundos ───────────────────────────────────────────────
        liquidez_map: dict[str, float] = {}
        for conta in set(list(rf_map.keys()) + list(fundos_liq_map.keys())):
            liquidez_map[conta] = (rf_map.get(conta) or 0) + (fundos_liq_map.get(conta) or 0)

        liquidez6_map: dict[str, float] = {}
        for conta in set(list(rf_map.keys()) + list(fundos_liq6_map.keys())):
            liquidez6_map[conta] = (rf_map.get(conta) or 0) + (fundos_liq6_map.get(conta) or 0)

        resultado = await repo.resumo_clientes_por_oferta(todos_clientes=todos_dict)

        for c in resultado["clientes"]:
            c["liquidez"] = liquidez_map.get(c["codigo_conta"], 0)
            c["liquidez6"] = liquidez6_map.get(c["codigo_conta"], 0)

        return resultado
