"""Importador completo do Diversificador XP — todas as classes de ativos.

Mapeamento de DSC_PRODUTO → tabela destino:
  Renda Fixa / Tesouro Direto / FIXED INCOME  → ass_renda_fixa (já existente)
  Renda Variável → Ação                        → ass_posicoes_rv (tipo=ACAO)
  Renda Variável → Fundo Imobiliário           → ass_posicoes_rv (tipo=FII)
  Renda Variável → Opção                       → ass_posicoes_rv (tipo=OPCAO)
  Renda Variável → Aluguel de Ações            → ass_posicoes_rv (tipo=ALUGUEL)
  Fundos (todos os sub)                        → ass_posicoes_fundos
  Previdência (todos os sub)                   → ass_posicoes_prev
  Somente Financeiro / OTHER PRODUCTS          → ignorado
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import structlog

from src.assistente.domain.entities.posicao_rv import PosicaoRV, _inferir_ticker
from src.assistente.domain.entities.posicao_fundo import PosicaoFundo, _inferir_tipo_fundo
from src.assistente.domain.entities.posicao_prev import PosicaoPrev, _inferir_tipo_prev

logger = structlog.get_logger(__name__)

_PRODUTOS_RF       = {"Renda Fixa", "Tesouro Direto", "FIXED INCOME"}
_PRODUTOS_RV       = {"Renda Variável"}
_PRODUTOS_FUNDOS   = {"Fundos"}
_PRODUTOS_PREV     = {"Previdência"}
_PRODUTOS_IGNORAR  = {"Somente Financeiro", "OTHER PRODUCTS"}

_SUB_RV = {
    "Ação": "ACAO",
    "Fundo Imobiliário": "FII",
    "Opção": "OPCAO",
    "Aluguel de Ações": "ALUGUEL",
}


@dataclass
class ResultadoImportacaoCompleta:
    posicoes_rv: list[PosicaoRV] = field(default_factory=list)
    posicoes_fundos: list[PosicaoFundo] = field(default_factory=list)
    posicoes_prev: list[PosicaoPrev] = field(default_factory=list)
    ignoradas: int = 0
    data_referencia: date | None = None


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return date.fromisoformat(str(value)[:10])
    except (ValueError, TypeError):
        return None


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _str_clean(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s.upper() not in ("NÃO APLICÁVEL", "NÃO ENCONTRADO", "") else None


def importar_todas_classes(
    caminho_diversificador: str | Path,
    nome_cliente_por_conta: dict[str, str] | None = None,
) -> ResultadoImportacaoCompleta:
    """Lê o Diversificador e retorna posições de TODAS as classes (exceto RF já importada).

    A Renda Fixa continua sendo importada pela função original em
    planilha_diversificador_importer.py. Esta função foca nas demais classes.

    Args:
        caminho_diversificador: Caminho para o arquivo .xlsx do Diversificador.
        nome_cliente_por_conta: Mapa {codigo_conta: nome} para enriquecer os registros.

    Returns:
        ResultadoImportacaoCompleta com listas separadas por classe.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl") from e

    logger.info("lendo_diversificador_completo", caminho=str(caminho_diversificador))
    wb = openpyxl.load_workbook(caminho_diversificador, read_only=True, data_only=True)
    aba = "Relatorio" if "Relatorio" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]

    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]  # type: ignore[arg-type]
    idx = {h: i for i, h in enumerate(header)}

    resultado = ResultadoImportacaoCompleta()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        produto     = str(row[idx.get("DSC_PRODUTO", 2)] or "").strip()
        sub_produto = str(row[idx.get("DSC_SUB_PRODUTO", 3)] or "").strip()
        conta       = str(row[idx.get("COD_CLIENTE", 1)] or "").strip()
        dsc_ativo   = str(row[idx.get("DSC_ATIVO", 5)] or "").strip()
        emissor     = _str_clean(row[idx.get("DSC_EMISSOR", 6)])
        cnpj_fundo  = _str_clean(row[idx.get("DSC_CNPJ_FUNDO", 4)])
        quantidade  = _parse_float(row[idx.get("VAL_QUANTIDADE", 8)])
        valor_net   = _parse_float(row[idx.get("VAL_NET", 9)])
        data_venc   = _parse_date(row[idx.get("DAT_DATA_VENCIMENTO", 7)])
        data_fato   = _parse_date(row[idx.get("DAT_DATA_FATO", 10)])
        nome        = (nome_cliente_por_conta or {}).get(conta)

        if resultado.data_referencia is None and data_fato:
            resultado.data_referencia = data_fato

        # ── Renda Fixa → já tratada pelo importer original ────────────────────
        if produto in _PRODUTOS_RF:
            resultado.ignoradas += 1
            continue

        # ── Ignorados ─────────────────────────────────────────────────────────
        if produto in _PRODUTOS_IGNORAR:
            resultado.ignoradas += 1
            continue

        # ── Renda Variável ────────────────────────────────────────────────────
        if produto in _PRODUTOS_RV:
            tipo_rv = _SUB_RV.get(sub_produto)
            if not tipo_rv:
                resultado.ignoradas += 1
                continue
            ticker = _inferir_ticker(dsc_ativo, emissor or "")
            resultado.posicoes_rv.append(
                PosicaoRV(
                    id=str(uuid.uuid4()),
                    codigo_conta=conta,
                    nome_cliente=nome,
                    tipo=tipo_rv,
                    ticker=ticker,
                    dsc_ativo=dsc_ativo,
                    emissor=emissor,
                    quantidade=quantidade,
                    valor_net=valor_net,
                    data_vencimento=data_venc,
                    data_referencia=data_fato,
                )
            )
            continue

        # ── Fundos ────────────────────────────────────────────────────────────
        if produto in _PRODUTOS_FUNDOS:
            tipo_fundo = _inferir_tipo_fundo(sub_produto)
            resultado.posicoes_fundos.append(
                PosicaoFundo(
                    id=str(uuid.uuid4()),
                    codigo_conta=conta,
                    nome_cliente=nome,
                    tipo_fundo=tipo_fundo,
                    cnpj_fundo=cnpj_fundo,
                    nome_fundo=dsc_ativo or sub_produto or None,
                    gestora=emissor,
                    valor_net=valor_net,
                    data_referencia=data_fato,
                )
            )
            continue

        # ── Previdência ───────────────────────────────────────────────────────
        if produto in _PRODUTOS_PREV:
            tipo_prev = _inferir_tipo_prev(sub_produto)
            resultado.posicoes_prev.append(
                PosicaoPrev(
                    id=str(uuid.uuid4()),
                    codigo_conta=conta,
                    nome_cliente=nome,
                    tipo_fundo=tipo_prev,
                    nome_fundo=dsc_ativo or sub_produto or None,
                    gestora=emissor,
                    valor_net=valor_net,
                    data_referencia=data_fato,
                )
            )
            continue

        resultado.ignoradas += 1

    wb.close()
    logger.info(
        "importacao_completa_concluida",
        rv=len(resultado.posicoes_rv),
        fundos=len(resultado.posicoes_fundos),
        prev=len(resultado.posicoes_prev),
        ignoradas=resultado.ignoradas,
    )
    return resultado
