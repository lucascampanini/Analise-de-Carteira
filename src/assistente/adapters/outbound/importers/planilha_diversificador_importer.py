"""Importador da planilha Diversificador XP.

Lê o arquivo Diversificacao-A{assessor}-Ref.{data}.xlsx e extrai
posições de renda fixa (Renda Fixa + Tesouro Direto) com vencimento definido.

Estrutura esperada (aba 'Relatorio'):
  DSC_ASSESSOR | COD_CLIENTE | DSC_PRODUTO | DSC_SUB_PRODUTO | DSC_CNPJ_FUNDO
  DSC_ATIVO    | DSC_EMISSOR | DAT_DATA_VENCIMENTO | VAL_QUANTIDADE | VAL_NET
  DAT_DATA_FATO
"""

from __future__ import annotations

import uuid
from datetime import date, datetime
from pathlib import Path

import structlog

from src.assistente.domain.entities.renda_fixa import (
    RendaFixa,
    _inferir_indexador,
    _inferir_tipo,
)

logger = structlog.get_logger(__name__)

# Produtos que possuem data de vencimento relevante
_PRODUTOS_RF = {"Renda Fixa", "Tesouro Direto"}


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return date.fromisoformat(str(value)[:10])
    except (ValueError, TypeError):
        return None


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _str_clean(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s.upper() not in ("NÃO APLICÁVEL", "NÃO ENCONTRADO", "") else None


def importar_renda_fixa(
    caminho_diversificador: str | Path,
    nome_cliente_por_conta: dict[str, str] | None = None,
) -> list[RendaFixa]:
    """Lê o Diversificador e retorna posições RF com vencimento.

    Args:
        caminho_diversificador: Caminho para o arquivo .xlsx do Diversificador.
        nome_cliente_por_conta: Mapa {codigo_conta: nome_cliente} para enriquecer
            os registros. Se None, nome_cliente fica em branco.

    Returns:
        Lista de RendaFixa prontos para persistência.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl") from e

    logger.info("lendo_diversificador", caminho=str(caminho_diversificador))
    wb = openpyxl.load_workbook(caminho_diversificador, read_only=True, data_only=True)

    aba = "Relatorio" if "Relatorio" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]

    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]  # type: ignore[arg-type]
    idx = {h: i for i, h in enumerate(header)}

    posicoes: list[RendaFixa] = []
    ignoradas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        produto = str(row[idx.get("DSC_PRODUTO", 2)] or "").strip()
        if produto not in _PRODUTOS_RF:
            ignoradas += 1
            continue

        vencimento = _parse_date(row[idx.get("DAT_DATA_VENCIMENTO", 7)])
        if not vencimento:
            ignoradas += 1
            continue

        conta = str(row[idx.get("COD_CLIENTE", 1)] or "").strip()
        sub_produto = str(row[idx.get("DSC_SUB_PRODUTO", 3)] or "").strip()
        dsc_ativo = str(row[idx.get("DSC_ATIVO", 5)] or "").strip()
        emissor = _str_clean(row[idx.get("DSC_EMISSOR", 6)])
        val_net = _parse_float(row[idx.get("VAL_NET", 9)])
        data_fato = _parse_date(row[idx.get("DAT_DATA_FATO", 10)])

        tipo = _inferir_tipo(produto, sub_produto, dsc_ativo)
        indexador = _inferir_indexador(dsc_ativo)
        nome_cliente = (nome_cliente_por_conta or {}).get(conta)

        posicoes.append(
            RendaFixa(
                id=str(uuid.uuid4()),
                codigo_conta=conta,
                nome_cliente=nome_cliente,
                tipo_ativo=tipo,
                dsc_ativo=dsc_ativo,
                emissor=emissor,
                indexador=indexador,
                data_vencimento=vencimento,
                valor_aplicado=val_net,
                data_referencia=data_fato,
            )
        )

    wb.close()
    logger.info(
        "importacao_diversificador_concluida",
        posicoes_rf=len(posicoes),
        ignoradas=ignoradas,
    )
    return posicoes
