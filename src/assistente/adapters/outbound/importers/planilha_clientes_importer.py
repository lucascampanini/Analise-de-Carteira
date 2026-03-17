"""Importador de clientes a partir das planilhas XP.

Cruza dois arquivos:
  1. RelatorioSaldoConsolidado — código conta, nome, saldos D0/D+1/D+2/D+3
  2. Positivador              — profissão, nascimento, cadastro, net, suitability

Filtro: apenas clientes do assessor 69567.
"""

from __future__ import annotations

import uuid
from datetime import date
from pathlib import Path

import structlog

from src.assistente.domain.entities.cliente_assessor import ClienteAssessor

logger = structlog.get_logger(__name__)


def _parse_date(value: object) -> date | None:
    """Converte valor Excel para date, tolerando formatos variados (ISO e DD/MM/YYYY)."""
    if value is None:
        return None
    from datetime import datetime as _dt
    if isinstance(value, _dt):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # DD/MM/YYYY (formato do Positivador XP)
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        try:
            return _dt.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            pass
    # YYYY-MM-DD (ISO)
    try:
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)  # type: ignore[arg-type]
    except (ValueError, TypeError):
        return None


def _str(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def importar_clientes(
    caminho_relatorio_saldo: str | Path,
    caminho_positivador: str | Path,
    codigo_assessor: str = "69567",
) -> list[ClienteAssessor]:
    """Lê as duas planilhas, filtra pelo assessor e retorna lista de ClienteAssessor.

    Args:
        caminho_relatorio_saldo: Caminho para o arquivo RelatorioSaldoConsolidado.xlsx
        caminho_positivador: Caminho para o arquivo Positivador.xlsx
        codigo_assessor: Código do assessor para filtrar (padrão: 69567)

    Returns:
        Lista de ClienteAssessor prontos para persistência.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl") from e

    # ── 1. Ler RelatorioSaldoConsolidado ──────────────────────────────────────
    logger.info("lendo_relatorio_saldo", caminho=str(caminho_relatorio_saldo))
    wb_saldo = openpyxl.load_workbook(caminho_relatorio_saldo, read_only=True, data_only=True)
    ws_saldo = wb_saldo.active

    saldo_por_conta: dict[str, dict[str, object]] = {}
    header_saldo = [str(c.value).strip() if c.value else "" for c in next(ws_saldo.iter_rows())]  # type: ignore[arg-type]

    idx = {h: i for i, h in enumerate(header_saldo)}
    for row in ws_saldo.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        assessor_val = str(row[idx.get("Assessor", 2)]).strip()
        if assessor_val != codigo_assessor:
            continue
        conta = str(row[idx["Conta"]]).strip()
        saldo_por_conta[conta] = {
            "nome": _str(row[idx.get("Cliente", 1)]),
            "d0": _parse_float(row[idx.get("D0", 3)]),
            "d1": _parse_float(row[idx.get("D+1", 4)]),
            "d2": _parse_float(row[idx.get("D+2", 5)]),
            "d3": _parse_float(row[idx.get("D+3", 6)]),
        }
    wb_saldo.close()
    logger.info("clientes_no_relatorio_saldo", total=len(saldo_por_conta))

    # ── 2. Ler Positivador ────────────────────────────────────────────────────
    logger.info("lendo_positivador", caminho=str(caminho_positivador))
    wb_pos = openpyxl.load_workbook(caminho_positivador, read_only=True, data_only=True)
    ws_pos = wb_pos.active

    header_pos = [str(c.value).strip() if c.value else "" for c in next(ws_pos.iter_rows())]  # type: ignore[arg-type]
    idx_pos = {h: i for i, h in enumerate(header_pos)}

    # Mapeamento flexível: tenta vários nomes possíveis para cada coluna
    def _col(*nomes: str, fallback: int) -> int:
        for n in nomes:
            if n in idx_pos:
                return idx_pos[n]
        # busca parcial insensível a acento/case
        for n in nomes:
            nl = n.lower()
            for k, v in idx_pos.items():
                if nl in k.lower():
                    return v
        return fallback

    col_conta     = _col("COD_CLIENTE", "Cliente",              fallback=1)
    col_profissao = _col("DSC_PROFISSAO", "Profissão", "Profissao", fallback=2)
    col_nasc      = _col("DAT_DATA_NASCIMENTO", "Data de Nascimento", "Nascimento", fallback=7)
    col_cadastro  = _col("DAT_DATA_CADASTRO", "Data de Cadastro", "Cadastro",   fallback=5)
    col_net       = _col("VAL_NET_EM_M", "Net Em M", "Net em M",               fallback=31)
    col_suitab    = _col("DSC_SUITABILITY", "Suitability", "Perfil",            fallback=6)
    col_segmento  = _col("DSC_SEGMENTO", "Segmento",                            fallback=4)

    logger.info("positivador_colunas_detectadas",
        conta=col_conta, profissao=col_profissao, nascimento=col_nasc,
        net=col_net, suitability=col_suitab)

    positivador_por_conta: dict[str, dict[str, object]] = {}
    for row in ws_pos.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        conta = str(row[col_conta]).strip()
        positivador_por_conta[conta] = {
            "profissao": _str(row[col_profissao]),
            "data_nascimento": _parse_date(row[col_nasc]),
            "data_cadastro": _parse_date(row[col_cadastro]),
            "net": _parse_float(row[col_net]),
            "suitability": _str(row[col_suitab]),
            "segmento": _str(row[col_segmento]),
        }
    wb_pos.close()
    logger.info("clientes_no_positivador", total=len(positivador_por_conta))

    # ── 3. Cruzar e montar entidades ──────────────────────────────────────────
    clientes: list[ClienteAssessor] = []
    sem_match = 0

    for conta, dados_saldo in saldo_por_conta.items():
        dados_pos = positivador_por_conta.get(conta, {})
        if not dados_pos:
            sem_match += 1

        clientes.append(
            ClienteAssessor(
                id=str(uuid.uuid4()),
                codigo_conta=conta,
                nome=str(dados_saldo.get("nome") or ""),
                profissao=str(dados_pos.get("profissao") or "") or None,
                data_nascimento=dados_pos.get("data_nascimento"),  # type: ignore[arg-type]
                data_cadastro=dados_pos.get("data_cadastro"),  # type: ignore[arg-type]
                net=dados_pos.get("net") or dados_saldo.get("d0"),  # type: ignore[arg-type]
                suitability=str(dados_pos.get("suitability") or "") or None,
                segmento=str(dados_pos.get("segmento") or "") or None,
                saldo_d0=dados_saldo.get("d0"),  # type: ignore[arg-type]
                saldo_d1=dados_saldo.get("d1"),  # type: ignore[arg-type]
                saldo_d2=dados_saldo.get("d2"),  # type: ignore[arg-type]
                saldo_d3=dados_saldo.get("d3"),  # type: ignore[arg-type]
            )
        )

    logger.info(
        "importacao_concluida",
        total_clientes=len(clientes),
        sem_match_positivador=sem_match,
    )
    return clientes
