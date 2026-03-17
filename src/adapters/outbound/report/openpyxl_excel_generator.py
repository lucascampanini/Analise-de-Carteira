"""Adapter: OpenpyxlExcelGenerator - workbook Excel de consolidação de carteiras."""

from __future__ import annotations

import io
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.application.dto.consolidacao_dto import ConsolidacaoDTO

# ── Paleta de cores ────────────────────────────────────────────────────────
_NAVY   = "1B3A6B"
_GOLD   = "C8A951"
_BLUE   = "2E5FA3"
_LBLUE  = "D6E4F0"
_GREEN  = "1E8449"
_LGREEN = "D5F5E3"
_RED    = "C0392B"
_LRED   = "FADBD8"
_GRAY   = "F2F3F4"
_WHITE  = "FFFFFF"
_DARK   = "2C3E50"
_LGRAY  = "BDC3C7"
_YELLOW = "FFF9C4"
_ORANGE = "F0A500"

_IDX_COLORS = {
    "CDI":   _LGREEN,
    "IPCA":  _LBLUE,
    "Pre":   _YELLOW,
    "PRE":   _YELLOW,
    "Multi": "FFE0B2",
    "MULTI": "FFE0B2",
    "RV":    "F8BBD9",
}

_BRL = "R$ #,##0.00"
_PCT = "0.00%"
_PCT0 = "0.0%"
_DT  = "DD/MM/YYYY"
_SIGN_PCT = "+0.00%;-0.00%"


class OpenpyxlExcelGenerator:
    """Gera workbook Excel de 6 abas para consolidação de carteiras.

    Implementa ExcelConsolidadoPort.
    """

    async def generate_excel(self, dto: ConsolidacaoDTO) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._build_premissas(wb, dto)
        self._build_posicao_consolidada(wb, dto)
        self._build_fluxo_caixa(wb, dto)
        self._build_projecao_anual(wb, dto)
        self._build_cenarios(wb, dto)
        self._build_acoes(wb, dto)

        # Cores das abas
        tab_colors = [_NAVY, _BLUE, _GREEN, _GOLD, _RED, "8E44AD"]
        for ws, cor in zip(wb.worksheets, tab_colors):
            ws.sheet_properties.tabColor = cor

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── ABA 1: PREMISSAS ──────────────────────────────────────────────────

    def _build_premissas(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("PREMISSAS")
        ws.sheet_view.showGridLines = False
        for i, w in enumerate([30, 14, 14, 14, 14, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        anos = dto.anos_projecao
        self._mhdr(ws, 1, 1, 7, "  PREMISSAS DA SIMULACAO", _NAVY, _WHITE, 14)
        self._mhdr(ws, 2, 1, 7,
                   f"Fonte: {dto.fonte_premissas} | CDI atual: {dto.cdi_atual_aa:.2f}% a.a. "
                   "| Valores em amarelo sao editaveis",
                   _BLUE, _WHITE, 9)

        # Tabela de taxas
        self._mhdr(ws, 4, 1, 7, "TAXAS PROJETADAS (% ao ano) — EDITE PARA SIMULAR", _BLUE, _WHITE, 11)
        for c, lbl in enumerate(["Indicador"] + [str(a) for a in anos], 1):
            self._hdr(ws, 5, c, lbl, _DARK, _WHITE, 10, center=(c > 1))

        # Encontrar premissas base (cenário BASE)
        base = next((c for c in dto.cenarios if "BASE" in c.nome), None)

        # Extrai taxas do DTO (usamos os valores do fallback pois DTO não expõe premissas diretamente)
        from src.domain.value_objects.premissas_mercado import PremissasMercado
        prem = PremissasMercado.fallback()

        for ridx, (lbl, attr) in enumerate([
            ("CDI (% a.a.)", "cdi_pct_aa"),
            ("IPCA (% a.a.)", "ipca_pct_aa"),
            ("IGP-M (% a.a.)", "igpm_pct_aa"),
            ("SELIC (% a.a.)", "selic_pct_aa"),
        ], 6):
            self._dc(ws, ridx, 1, lbl, bold=True, bg=_GRAY)
            for cidx, ano in enumerate(anos, 2):
                val = getattr(prem.para_ano(ano), attr) / 100
                c = self._dc(ws, ridx, cidx, val, fmt=_PCT, center=True, bg=_YELLOW)
                c.font = Font(bold=True, color=_DARK, size=10, name="Calibri")

        # Fundos estimados
        self._mhdr(ws, 11, 1, 7, "PREMISSAS PARA FUNDOS SEM BENCHMARK DEFINIDO", _BLUE, _WHITE, 11)
        self._hdr(ws, 12, 1, "Fundo", _DARK, _WHITE, 10)
        self._hdr(ws, 12, 2, "Retorno Projetado", _DARK, _WHITE, 10, center=True)
        ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=7)

        fundos_est = [(a.nome, "CDI + 2,00% a.a.") for a in dto.ativos_rf if a.indexador == "Multi"]
        fundos_est += [(a.nome, "CDI + 3,00% a.a. (Long Biased)") for a in dto.ativos_rf if a.indexador == "RV"]
        for ridx, (nm, prem_txt) in enumerate(fundos_est, 13):
            self._dc(ws, ridx, 1, nm, bg=_GRAY)
            c = self._dc(ws, ridx, 2, prem_txt, bg=_YELLOW, center=True, bold=True)
            ws.merge_cells(start_row=ridx, start_column=2, end_row=ridx, end_column=7)

        # Tabela IR
        r_ir = 13 + len(fundos_est) + 2
        self._mhdr(ws, r_ir, 1, 7, "TABELA REGRESSIVA DE IR — Renda Fixa e Fundos", _BLUE, _WHITE, 11)
        self._hdr(ws, r_ir + 1, 1, "Prazo da Aplicacao", _DARK, _WHITE, 10)
        self._hdr(ws, r_ir + 1, 2, "Aliquota IR", _DARK, _WHITE, 10, center=True)
        self._hdr(ws, r_ir + 1, 3, "Observacao", _DARK, _WHITE, 10)
        ws.merge_cells(start_row=r_ir + 1, start_column=3, end_row=r_ir + 1, end_column=7)
        for ridx, (pr, al, obs) in enumerate([
            ("Ate 180 dias", "22,50%", ""),
            ("181 a 360 dias", "20,00%", ""),
            ("361 a 720 dias", "17,50%", ""),
            ("Acima de 720 dias", "15,00%", "Aliquota aplicada na maioria das posicoes"),
            ("CRI / CRA / Debentures incent.", "ISENTO", "Beneficio fiscal para Pessoa Fisica"),
        ], r_ir + 2):
            self._dc(ws, ridx, 1, pr, bg=_GRAY)
            self._dc(ws, ridx, 2, al, center=True, bold=True,
                     bg=_LGREEN if "ISENTO" in al else _YELLOW)
            self._dc(ws, ridx, 3, obs, italic=True, bg=_GRAY)
            ws.merge_cells(start_row=ridx, start_column=3, end_row=ridx, end_column=7)

        self._mhdr(ws, r_ir + 8, 1, 7,
                   "Reinvestimento: 100% CDI (liquidez diaria). Acoes sem projecao de preco.",
                   _GOLD, "1B3A6B", 10)

    # ── ABA 2: POSICAO CONSOLIDADA ────────────────────────────────────────

    def _build_posicao_consolidada(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("POSICAO CONSOLIDADA")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A7"
        for i, w in enumerate([12, 40, 16, 10, 12, 14, 14, 16, 10, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        self._mhdr(ws, 1, 1, 10,
                   f"  POSICAO CONSOLIDADA | Data-base: {dto.data_referencia}", _NAVY, _WHITE, 14)

        # Patrimônio total
        self._dc(ws, 2, 1, "PATRIMONIO TOTAL:", bold=True, bg=_LBLUE, fg=_NAVY, sz=11)
        self._dc(ws, 2, 2, dto.total_geral, bold=True, bg=_LBLUE, fg=_NAVY, sz=12, fmt=_BRL)
        for ci in range(3, 11):
            self._dc(ws, 2, ci, "", bg=_LBLUE)

        # Totais por conta
        contas = sorted({a.conta for a in dto.ativos_rf})
        for acc_idx, acc in enumerate(contas):
            tot_conta = sum(a.posicao for a in dto.ativos_rf if a.conta == acc)
            c1 = 1 + acc_idx * 3
            self._dc(ws, 3, c1, f"Conta {acc}:", bold=True, bg=_GRAY)
            self._dc(ws, 3, c1 + 1, tot_conta, bold=True, bg=_GRAY, fmt=_BRL)

        # Resumo por indexador
        self._mhdr(ws, 5, 1, 10, "RESUMO POR INDEXADOR", _BLUE, _WHITE, 11)
        for c, lbl in enumerate(["Indexador", "Posicao (R$)", "% Total"], 1):
            self._hdr(ws, 6, c, lbl, _DARK, _WHITE, 10, center=(c > 1))
        for r_off, res in enumerate(dto.resumo_por_indexador, 7):
            bg = _IDX_COLORS.get(res.indexador, _GRAY)
            lbl = "Pre-Fixado" if res.indexador in ("Pre", "PRE") else res.indexador
            self._dc(ws, r_off, 1, lbl, bold=True, bg=bg, center=True)
            self._dc(ws, r_off, 2, res.total, fmt=_BRL, bg=bg)
            self._dc(ws, r_off, 3, res.percentual, fmt=_PCT0, bg=bg, center=True)

        row_total = 7 + len(dto.resumo_por_indexador)
        self._dc(ws, row_total, 1, "TOTAL", bold=True, bg=_NAVY, fg=_WHITE, center=True)
        self._dc(ws, row_total, 2, dto.total_geral, bold=True, bg=_NAVY, fg=_WHITE, fmt=_BRL)
        self._dc(ws, row_total, 3, 1.0, bold=True, bg=_NAVY, fg=_WHITE, fmt=_PCT0, center=True)

        # Tabela detalhada
        row = row_total + 2
        for c, h in enumerate(["Conta", "Ativo / Fundo", "Tipo", "Indexador", "Taxa",
                                "Aplicacao", "Vencimento", "Posicao", "IR Isento", "Obs."], 1):
            self._hdr(ws, row, c, h, _NAVY, _WHITE, 9, center=True, wrap=True)
        row += 1

        prev_conta = None
        for a in dto.ativos_rf:
            if a.conta != prev_conta:
                self._mhdr(ws, row, 1, 10, f"  CONTA {a.conta}", _BLUE, _WHITE, 10)
                row += 1
                prev_conta = a.conta
            bg = _IDX_COLORS.get(a.indexador, _GRAY)
            aplica = a.data_aplicacao or "—"
            venc = a.data_vencimento or "Aberto"
            self._dc(ws, row, 1, a.conta, center=True, bg=bg)
            self._dc(ws, row, 2, a.nome, bg=bg)
            self._dc(ws, row, 3, a.tipo, bg=bg, center=True)
            self._dc(ws, row, 4, "Pre-Fixado" if a.indexador in ("Pre","PRE") else a.indexador,
                     bg=bg, center=True, bold=True)
            self._dc(ws, row, 5, a.taxa_formatada, bg=bg, center=True)
            self._dc(ws, row, 6, aplica, center=True, bg=bg)
            self._dc(ws, row, 7, venc, center=True, bg=bg)
            self._dc(ws, row, 8, a.posicao, fmt=_BRL, bg=bg)
            self._dc(ws, row, 9, "SIM" if a.ir_isento else "NAO", center=True,
                     bg=_LGREEN if a.ir_isento else _LRED,
                     fg=_GREEN if a.ir_isento else _RED, bold=True)
            self._dc(ws, row, 10, a.nota, italic=True, bg=bg)
            row += 1

        # Total RF+Fundos
        self._mhdr(ws, row, 1, 7, "SUBTOTAL RF + FUNDOS", _DARK, _WHITE, 10)
        self._dc(ws, row, 8, dto.total_rf_fundos, bold=True, fmt=_BRL, bg=_DARK, fg=_WHITE)
        row += 1

        # Ações
        if dto.acoes:
            self._mhdr(ws, row, 1, 10, "  ACOES — RENDA VARIAVEL", "E8EAF6", _DARK, 10)
            row += 1
            for a in dto.acoes:
                self._dc(ws, row, 1, a.conta, center=True, bg="E8EAF6")
                self._dc(ws, row, 2, f"{a.ticker} - {a.nome}", bg="E8EAF6")
                self._dc(ws, row, 3, "Acao", bg="E8EAF6", center=True)
                self._dc(ws, row, 4, "RV", bg="E8EAF6", center=True, bold=True)
                self._dc(ws, row, 5, f"R${a.ultimo_preco:.2f}", bg="E8EAF6", center=True)
                self._dc(ws, row, 6, "—", center=True, bg="E8EAF6")
                self._dc(ws, row, 7, "Aberto", center=True, bg="E8EAF6")
                self._dc(ws, row, 8, a.posicao, fmt=_BRL, bg="E8EAF6")
                self._dc(ws, row, 9, "N/A", center=True, bg="E8EAF6")
                self._dc(ws, row, 10, f"Qtd: {a.qtd:,}", bg="E8EAF6")
                row += 1

        self._mhdr(ws, row, 1, 7, "TOTAL GERAL (RF + Fundos + Acoes)", _NAVY, _WHITE, 11)
        self._dc(ws, row, 8, dto.total_geral, bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE, sz=11)

    # ── ABA 3: FLUXO DE CAIXA RF ─────────────────────────────────────────

    def _build_fluxo_caixa(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("FLUXO DE CAIXA RF")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"
        for i, w in enumerate([12, 42, 12, 18, 16, 10, 16, 10, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        self._mhdr(ws, 1, 1, 9,
                   "  FLUXO DE CAIXA RF | Juros, Amortizacoes e Reinvestimento", _NAVY, _WHITE, 13)
        self._mhdr(ws, 2, 1, 9,
                   "Cupons projetados com taxas Focus. Recebimentos reinvestidos a 100% CDI.",
                   _BLUE, _WHITE, 9)
        for c, h in enumerate(["Data", "Ativo", "Conta", "Evento", "Bruto (R$)",
                                "IR (%)", "Liquido (R$)", "Isento IR", "Acum. Reinvest. CDI"], 1):
            self._hdr(ws, 4, c, h, _NAVY, _WHITE, 9, center=True, wrap=True)

        row = 5
        row_colors = [_GRAY, _WHITE]
        prev_yr = None
        for i, cf in enumerate(dto.fluxos_caixa):
            yr = cf.data[:4]
            if yr != prev_yr:
                self._mhdr(ws, row, 1, 9, f"  {yr}", _BLUE, _WHITE, 10)
                row += 1
                prev_yr = yr

            bg = row_colors[i % 2]
            self._dc(ws, row, 1, cf.data, center=True, bg=bg)
            self._dc(ws, row, 2, cf.ativo, bg=bg)
            self._dc(ws, row, 3, cf.conta, center=True, bg=bg)
            self._dc(ws, row, 4, cf.evento, center=True, bg=bg)
            self._dc(ws, row, 5, cf.valor_bruto, fmt=_BRL, bg=bg)
            self._dc(ws, row, 6, cf.aliquota_ir, fmt=_PCT, center=True,
                     bg=_LGREEN if cf.ir_isento else bg)
            self._dc(ws, row, 7, cf.valor_liquido, fmt=_BRL, bg=bg)
            self._dc(ws, row, 8, "SIM" if cf.ir_isento else "NAO", center=True,
                     bg=_LGREEN if cf.ir_isento else _LRED,
                     fg=_GREEN if cf.ir_isento else _RED, bold=True)
            self._dc(ws, row, 9, cf.acumulado_reinvest, fmt=_BRL, bg=_LGREEN)
            row += 1

        # Totais
        self._mhdr(ws, row, 1, 4, "TOTAIS", _NAVY, _WHITE, 10)
        self._dc(ws, row, 5, sum(c.valor_bruto for c in dto.fluxos_caixa),
                 bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE)
        self._dc(ws, row, 7, sum(c.valor_liquido for c in dto.fluxos_caixa),
                 bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE)
        if dto.fluxos_caixa:
            self._dc(ws, row, 9, dto.fluxos_caixa[-1].acumulado_reinvest,
                     bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE)

    # ── ABA 4: PROJECAO ANUAL ─────────────────────────────────────────────

    def _build_projecao_anual(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("PROJECAO ANUAL")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C7"
        anos = dto.anos_projecao
        n = len(anos)
        for i, w in enumerate([12, 40, 14] + [16] * n, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        self._mhdr(ws, 1, 1, 3 + n,
                   "  PROJECAO ANUAL DE PATRIMONIO | RF + Fundos", _NAVY, _WHITE, 14)
        self._mhdr(ws, 2, 1, 3 + n,
                   "Projecao com taxas Focus BCB | Cupons reinvestidos a 100% CDI | Acoes em aba separada",
                   _BLUE, _WHITE, 9)

        # Header anos
        for c, h in enumerate(["Conta", "Ativo", f"Mar/{str(dto.data_referencia[:4])[2:]}"]
                               + [f"Dez/{str(a)[2:]}" for a in anos], 1):
            self._hdr(ws, 5, c, h, _NAVY, _WHITE, 9, center=True, wrap=True)

        # Linhas por ativo
        row = 6
        prev_conta = None
        for a in dto.ativos_rf:
            if a.conta != prev_conta:
                self._mhdr(ws, row, 1, 3 + n, f"  CONTA {a.conta}", _BLUE, _WHITE, 10)
                row += 1
                prev_conta = a.conta

            bg = _IDX_COLORS.get(a.indexador, _GRAY)
            self._dc(ws, row, 1, a.conta, center=True, bg=bg)
            self._dc(ws, row, 2, a.nome, bg=bg)
            self._dc(ws, row, 3, a.posicao, fmt=_BRL, bg=bg)

            # Para cada ano, só mostramos o total RF+Fundos no nível de ativo
            # (as projeções granulares ficam no handler — aqui mostramos apenas posicao e evolucao)
            for ci, yr in enumerate(anos, 4):
                venc = a.data_vencimento
                if venc and date.fromisoformat(venc).year < yr:
                    self._dc(ws, row, ci, "VENCIDO", center=True, bg=_LGRAY,
                              fg="999999", italic=True)
                else:
                    # Busca projeção do DTO
                    proj_ano = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
                    # Mostra só o valor da posicao (projeção individual não está no DTO por ativo)
                    self._dc(ws, row, ci, "—", center=True, bg=bg, italic=True, fg="888888")
            row += 1

        # Linha de reinvestimento
        self._dc(ws, row, 1, "", bg=_LGREEN)
        self._dc(ws, row, 2, "Reinvestimento de Cupons/Vencimentos (100% CDI)",
                 bold=True, bg=_LGREEN, fg=_GREEN)
        self._dc(ws, row, 3, 0, fmt=_BRL, bg=_LGREEN)
        for ci, yr in enumerate(anos, 4):
            proj = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
            val = proj.reinvestimento_acumulado if proj else 0.0
            self._dc(ws, row, ci, val, fmt=_BRL, bg=_LGREEN, bold=True)
        row += 1

        # Total RF+Fundos
        self._mhdr(ws, row, 1, 2, "TOTAL RF + FUNDOS (incl. reinvestimento)", _NAVY, _WHITE, 11)
        self._dc(ws, row, 3, dto.total_rf_fundos, bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE, sz=11)
        for ci, yr in enumerate(anos, 4):
            proj = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
            val = proj.total_rf_fundos if proj else 0.0
            self._dc(ws, row, ci, val, bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE, sz=11)
        row += 1

        # Ações (sem projeção)
        self._mhdr(ws, row, 1, 2, "Acoes (valor atual — sem projecao de preco)", "E8EAF6", _DARK, 10)
        self._dc(ws, row, 3, dto.total_acoes, fmt=_BRL, bg="E8EAF6")
        for ci in range(4, 4 + n):
            self._dc(ws, row, ci, "sem proj.", center=True, bg="E8EAF6", italic=True, fg="888888")
        row += 1

        # Total Geral
        self._mhdr(ws, row, 1, 2, "TOTAL GERAL (RF + Fundos + Acoes)", _GOLD, "1B3A6B", 12)
        self._dc(ws, row, 3, dto.total_geral, bold=True, fmt=_BRL, bg=_GOLD, fg=_NAVY, sz=12)
        for ci, yr in enumerate(anos, 4):
            proj = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
            val = proj.total_geral if proj else dto.total_geral
            self._dc(ws, row, ci, val, bold=True, fmt=_BRL, bg=_GOLD, fg=_NAVY, sz=12)
        row += 2

        # Variação e CAGR
        self._mhdr(ws, row, 1, 2, "Variacao vs. Hoje (RF+Fundos)", _GREEN, _WHITE, 10)
        self._dc(ws, row, 3, "Base", center=True, bg=_LGREEN, bold=True)
        for ci, yr in enumerate(anos, 4):
            proj = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
            var = proj.variacao_vs_hoje if proj else 0.0
            self._dc(ws, row, ci, var, fmt=_PCT, center=True,
                     bg=_LGREEN if var >= 0 else _LRED,
                     fg=_GREEN if var >= 0 else _RED, bold=True)
        row += 1

        self._mhdr(ws, row, 1, 2, "Retorno Medio Anualizado (CAGR)", _BLUE, _WHITE, 10)
        self._dc(ws, row, 3, "—", center=True, bg=_LBLUE)
        for ci, yr in enumerate(anos, 4):
            proj = next((p for p in dto.projecao_por_ano if p.ano == yr), None)
            cagr = proj.retorno_medio_aa if proj else 0.0
            self._dc(ws, row, ci, cagr, fmt=_PCT, center=True, bg=_LBLUE, bold=True)

    # ── ABA 5: CENARIOS ───────────────────────────────────────────────────

    def _build_cenarios(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("CENARIOS")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 32
        n_cen = len(dto.cenarios)
        for i in range(2, 2 + n_cen):
            ws.column_dimensions[get_column_letter(i)].width = 20

        self._mhdr(ws, 1, 1, 1 + n_cen, "  ANALISE DE CENARIOS | 5 Anos", _NAVY, _WHITE, 14)
        self._mhdr(ws, 2, 1, 1 + n_cen,
                   "Ajuste CDI/IPCA e veja o impacto no patrimonio. Valores em R$.",
                   _BLUE, _WHITE, 9)

        # Cores por cenário
        cen_colors = [_LBLUE, _LGREEN, _LRED, "FFE0B2"]

        self._hdr(ws, 4, 1, "Parametro", _DARK, _WHITE, 10)
        for ci, (cen, cor) in enumerate(zip(dto.cenarios, cen_colors), 2):
            self._hdr(ws, 4, ci, cen.nome, _DARK, _WHITE, 9, center=True, wrap=True)

        self._dc(ws, 5, 1, "Ajuste CDI (p.p.)", bold=True, bg=_GRAY)
        for ci, (cen, cor) in enumerate(zip(dto.cenarios, cen_colors), 2):
            self._dc(ws, 5, ci, cen.delta_cdi / 100, fmt=_SIGN_PCT, center=True, bg=cor, bold=True)

        self._dc(ws, 6, 1, "Ajuste IPCA (p.p.)", bold=True, bg=_GRAY)
        for ci, (cen, cor) in enumerate(zip(dto.cenarios, cen_colors), 2):
            self._dc(ws, 6, ci, cen.delta_ipca / 100, fmt=_SIGN_PCT, center=True, bg=cor, bold=True)

        self._mhdr(ws, 8, 1, 1 + n_cen,
                   "PATRIMONIO PROJETADO POR CENARIO (RF + Fundos, sem Acoes)", _BLUE, _WHITE, 11)
        self._hdr(ws, 9, 1, "Ano", _DARK, _WHITE, 10)
        for ci, cen in enumerate(dto.cenarios, 2):
            self._hdr(ws, 9, ci, cen.nome, _DARK, _WHITE, 9, center=True)

        for r_off, ano in enumerate(dto.anos_projecao, 10):
            self._dc(ws, r_off, 1, str(ano), bold=True, bg=_GRAY, center=True)
            for ci, (cen, cor) in enumerate(zip(dto.cenarios, cen_colors), 2):
                val = cen.patrimonio_por_ano.get(ano, 0.0)
                self._dc(ws, r_off, ci, val, fmt=_BRL, bg=cor,
                         bold=(ano == max(dto.anos_projecao)))

        r_sum = 10 + len(dto.anos_projecao) + 1
        self._mhdr(ws, r_sum, 1, 1 + n_cen,
                   "RETORNO TOTAL — PERIODO COMPLETO (vs. posicao atual RF+Fundos)",
                   _DARK, _WHITE, 11)
        for lbl, r_off, attr, fmt in [
            ("Patrimonio Final", 1, "patrimonio_final", _BRL),
            ("Ganho Absoluto (R$)", 2, "ganho_absoluto", _BRL),
            ("Retorno Total", 3, "retorno_total", _PCT),
            ("Retorno Anualizado (CAGR)", 4, "cagr", _PCT),
        ]:
            self._dc(ws, r_sum + r_off, 1, lbl, bold=True, bg=_GRAY)
            for ci, (cen, cor) in enumerate(zip(dto.cenarios, cen_colors), 2):
                self._dc(ws, r_sum + r_off, ci, getattr(cen, attr),
                         fmt=fmt, bg=cor, bold=True, center=(fmt != _BRL))

    # ── ABA 6: ACOES ─────────────────────────────────────────────────────

    def _build_acoes(self, wb, dto: ConsolidacaoDTO) -> None:
        ws = wb.create_sheet("ACOES")
        ws.sheet_view.showGridLines = False
        for i, w in enumerate([10, 28, 12, 14, 16, 14, 14, 28], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        self._mhdr(ws, 1, 1, 8,
                   f"  CARTEIRA DE ACOES | Data-base: {dto.data_referencia}",
                   _NAVY, _WHITE, 13)
        self._mhdr(ws, 2, 1, 8,
                   "Acoes nao incluidas na projecao quantitativa. Valor atual de mercado.",
                   _BLUE, _WHITE, 9)

        for c, h in enumerate(["Ticker", "Nome", "Qtd", "Ult. Cotacao",
                                "Posicao", "% Carteira", "Conta", "Obs."], 1):
            self._hdr(ws, 4, c, h, _NAVY, _WHITE, 9, center=True, wrap=True)

        for r_off, a in enumerate(dto.acoes, 5):
            bg = "F0F4FF"
            self._dc(ws, r_off, 1, a.ticker, bold=True, center=True, bg=bg)
            self._dc(ws, r_off, 2, a.nome, bg=bg)
            self._dc(ws, r_off, 3, a.qtd, fmt="#,##0", center=True, bg=bg)
            self._dc(ws, r_off, 4, a.ultimo_preco, fmt="R$ #,##0.00", center=True, bg=bg)
            self._dc(ws, r_off, 5, a.posicao, fmt=_BRL, bg=bg)
            self._dc(ws, r_off, 6, a.percentual_carteira, fmt=_PCT0, center=True, bg=bg)
            self._dc(ws, r_off, 7, a.conta, center=True, bg=bg)
            self._dc(ws, r_off, 8, "", bg=bg)

        row_tot = 5 + len(dto.acoes)
        self._mhdr(ws, row_tot, 1, 4, "TOTAL ACOES", _NAVY, _WHITE, 11)
        self._dc(ws, row_tot, 5, dto.total_acoes, bold=True, fmt=_BRL, bg=_NAVY, fg=_WHITE, sz=11)
        self._dc(ws, row_tot, 6, 1.0, bold=True, fmt=_PCT0, bg=_NAVY, fg=_WHITE, center=True)

    # ── Helpers de formatação ─────────────────────────────────────────────

    def _hdr(self, ws: Worksheet, r: int, c: int, v: Any,
             bg: str = _NAVY, fg: str = _WHITE, sz: int = 11,
             bold: bool = True, center: bool = True, wrap: bool = False) -> Any:
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap,
        )
        s = Side(style="thin", color=_LGRAY)
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def _dc(self, ws: Worksheet, r: int, c: int, v: Any,
            bold: bool = False, bg: str | None = None, fg: str = _DARK,
            sz: int = 10, fmt: str | None = None, center: bool = False,
            italic: bool = False, wrap: bool = False) -> Any:
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, color=fg, size=sz, name="Calibri", italic=italic)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap,
        )
        if fmt:
            cell.number_format = fmt
        s = Side(style="thin", color=_LGRAY)
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        return cell

    def _mhdr(self, ws: Worksheet, r: int, c1: int, c2: int, v: Any,
              bg: str = _NAVY, fg: str = _WHITE, sz: int = 12,
              bold: bool = True) -> Any:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=v)
        cell.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell
