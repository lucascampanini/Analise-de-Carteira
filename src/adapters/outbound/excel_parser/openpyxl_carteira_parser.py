"""Adapter: OpenpyxlCarteiraParser — lê planilha Excel e retorna PosicaoParsedDTO."""

from __future__ import annotations

import io
import logging
from decimal import Decimal, InvalidOperation

import openpyxl

from src.application.ports.outbound.pdf_parser_port import PosicaoParsedDTO

logger = logging.getLogger(__name__)

# Colunas obrigatórias (case-insensitive, strip de espaços)
COLUNAS_OBRIGATORIAS = {"ticker", "nome", "quantidade", "preco_medio", "valor_atual", "classe_ativo"}

# Colunas opcionais RF
COLUNAS_RF = {
    "subtipo_rf", "indexador_rf", "taxa_rf", "data_vencimento_rf",
    "data_carencia_rf", "liquidez_rf", "cnpj_emissor_rf",
    "rating_escala_rf", "rating_agencia_rf", "garantias_rf",
}


class OpenpyxlCarteiraParser:
    """Lê um .xlsx com posições de carteira e converte em PosicaoParsedDTO.

    Formato esperado da planilha (sheet "Carteira" ou primeira sheet):
    - Linha 1: cabeçalhos (nomes de coluna, case-insensitive)
    - Linha 2+: dados de cada posição

    Colunas obrigatórias:
        ticker, nome, quantidade, preco_medio, valor_atual, classe_ativo

    Colunas opcionais (setor, emissor, campos RF):
        setor, emissor, subtipo_rf, indexador_rf, taxa_rf,
        data_vencimento_rf, data_carencia_rf, liquidez_rf,
        cnpj_emissor_rf, rating_escala_rf, rating_agencia_rf, garantias_rf
    """

    def parse_carteira(self, excel_bytes: bytes) -> list[PosicaoParsedDTO]:
        """Extrai posições de um arquivo Excel de carteira.

        Args:
            excel_bytes: Conteúdo binário do arquivo .xlsx.

        Returns:
            Lista de posições parseadas.

        Raises:
            ValueError: Se o arquivo for inválido ou faltar colunas obrigatórias.
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Não foi possível abrir o arquivo Excel: {exc}") from exc

        # Prefer sheet "Carteira", fallback para primeira
        sheet_name = "Carteira" if "Carteira" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError("Planilha vazia ou sem dados (mínimo: 1 linha de cabeçalho + 1 linha de dados).")

        # Mapear cabeçalhos → índice de coluna
        header_row = rows[0]
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                col_map[str(cell).strip().lower()] = idx

        ausentes = COLUNAS_OBRIGATORIAS - col_map.keys()
        if ausentes:
            raise ValueError(
                f"Colunas obrigatórias ausentes na planilha: {sorted(ausentes)}. "
                f"Colunas encontradas: {sorted(col_map.keys())}"
            )

        posicoes: list[PosicaoParsedDTO] = []
        for row_num, row in enumerate(rows[1:], start=2):
            ticker = self._cell_str(row, col_map, "ticker")
            if not ticker:
                logger.debug("Linha %d ignorada: ticker vazio.", row_num)
                continue

            try:
                quantidade = self._cell_decimal(row, col_map, "quantidade")
                preco_medio = self._cell_float(row, col_map, "preco_medio")
                valor_atual = self._cell_float(row, col_map, "valor_atual")
            except ValueError as exc:
                raise ValueError(f"Linha {row_num} ({ticker}): {exc}") from exc

            dto = PosicaoParsedDTO(
                ticker=ticker.upper(),
                nome=self._cell_str(row, col_map, "nome") or ticker,
                quantidade=quantidade,
                preco_medio=preco_medio,
                valor_atual=valor_atual,
                classe_ativo=self._cell_str(row, col_map, "classe_ativo") or "ACAO",
                setor=self._cell_str(row, col_map, "setor") or "Não classificado",
                emissor=self._cell_str(row, col_map, "emissor") or ticker,
                # RF opcionais
                subtipo_rf=self._cell_str(row, col_map, "subtipo_rf"),
                indexador_rf=self._cell_str(row, col_map, "indexador_rf"),
                taxa_rf=self._cell_float_or_none(row, col_map, "taxa_rf"),
                data_vencimento_rf=self._cell_date_str(row, col_map, "data_vencimento_rf"),
                data_carencia_rf=self._cell_date_str(row, col_map, "data_carencia_rf"),
                liquidez_rf=self._cell_str(row, col_map, "liquidez_rf"),
                cnpj_emissor_rf=self._cell_str(row, col_map, "cnpj_emissor_rf"),
                rating_escala_rf=self._cell_str(row, col_map, "rating_escala_rf"),
                rating_agencia_rf=self._cell_str(row, col_map, "rating_agencia_rf"),
                garantias_rf=self._cell_str(row, col_map, "garantias_rf"),
            )
            posicoes.append(dto)
            logger.debug("Linha %d: %s parseado com sucesso.", row_num, ticker)

        if not posicoes:
            raise ValueError("Nenhuma posição válida encontrada na planilha.")

        logger.info("Excel parseado: %d posições carregadas.", len(posicoes))
        return posicoes

    # ── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _get(row: tuple, col_map: dict[str, int], col: str):
        idx = col_map.get(col)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _cell_str(self, row: tuple, col_map: dict[str, int], col: str) -> str | None:
        val = self._get(row, col_map, col)
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def _cell_decimal(self, row: tuple, col_map: dict[str, int], col: str) -> Decimal:
        val = self._get(row, col_map, col)
        if val is None:
            raise ValueError(f"Coluna '{col}' é obrigatória e está vazia.")
        try:
            return Decimal(str(val).replace(",", "."))
        except InvalidOperation:
            raise ValueError(f"Valor inválido em '{col}': '{val}' (esperado número).")

    def _cell_float(self, row: tuple, col_map: dict[str, int], col: str) -> float:
        val = self._get(row, col_map, col)
        if val is None:
            raise ValueError(f"Coluna '{col}' é obrigatória e está vazia.")
        try:
            return float(str(val).replace(",", "."))
        except ValueError:
            raise ValueError(f"Valor inválido em '{col}': '{val}' (esperado número).")

    def _cell_float_or_none(self, row: tuple, col_map: dict[str, int], col: str) -> float | None:
        val = self._get(row, col_map, col)
        if val is None or str(val).strip() == "":
            return None
        try:
            return float(str(val).replace(",", "."))
        except ValueError:
            return None

    def _cell_date_str(self, row: tuple, col_map: dict[str, int], col: str) -> str | None:
        """Retorna data no formato YYYY-MM-DD ou None."""
        from datetime import date, datetime
        val = self._get(row, col_map, col)
        if val is None or str(val).strip() == "":
            return None
        # openpyxl pode retornar datetime diretamente de células de data
        if isinstance(val, (datetime, date)):
            if isinstance(val, datetime):
                return val.date().isoformat()
            return val.isoformat()
        # Tenta parse de string
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        logger.warning("Não foi possível parsear data '%s' na coluna '%s'.", s, col)
        return None
