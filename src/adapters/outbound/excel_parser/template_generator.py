"""Gerador do template Excel para input manual de carteira."""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def gerar_template_excel() -> bytes:
    """Gera o arquivo .xlsx com template de carteira para preenchimento.

    Returns:
        Bytes do arquivo Excel (.xlsx).
    """
    wb = openpyxl.Workbook()

    _criar_sheet_carteira(wb)
    _criar_sheet_instrucoes(wb)
    _criar_sheet_valores_validos(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Sheet principal ───────────────────────────────────────────────────────────

COLUNAS = [
    # (nome_coluna, largura, obrigatorio, descricao)
    ("ticker",              18, True,  "Ticker do ativo (ex: PETR4, TESOURO_IPCA_2029, CDB_BANCO_X)"),
    ("nome",                35, True,  "Nome completo do ativo"),
    ("quantidade",          15, True,  "Quantidade (cotas, unidades ou valor nominal RF)"),
    ("preco_medio",         15, True,  "Preço médio de compra em R$"),
    ("valor_atual",         15, True,  "Valor atual da posição em R$"),
    ("classe_ativo",        20, True,  "ACAO | FII | ETF | RENDA_FIXA | BDR | FUNDO | CRIPTO"),
    ("setor",               22, False, "Setor (ex: Energia, Financeiro, Saúde)"),
    ("emissor",             25, False, "Emissor/empresa (ex: Petrobras, Itaú)"),
    # RF opcionais
    ("subtipo_rf",          20, False, "CDB | LCI | LCA | LIG | CRI | CRA | DEBENTURE | TESOURO_SELIC | TESOURO_IPCA | TESOURO_PREFIXADO | OUTRO"),
    ("indexador_rf",        18, False, "CDI_PERCENTUAL | CDI_MAIS | SELIC | IPCA_MAIS | PREFIXADO | IGPM_MAIS | TR_MAIS | OUTRO"),
    ("taxa_rf",             12, False, "Taxa numérica (ex: 12.5 para 12,5% ou 110 para 110% CDI)"),
    ("data_vencimento_rf",  20, False, "Data de vencimento DD/MM/AAAA (ex: 31/12/2027)"),
    ("data_carencia_rf",    20, False, "Data de carência DD/MM/AAAA (ou deixar vazio)"),
    ("liquidez_rf",         18, False, "DIARIA | NO_VENCIMENTO | D_MAIS_30 | D_MAIS_60 | D_MAIS_90"),
    ("cnpj_emissor_rf",     20, False, "CNPJ do emissor sem formatação (14 dígitos)"),
    ("rating_escala_rf",    18, False, "AAA | AA_MAIS | AA | AA_MENOS | A_MAIS | A | BBB | BB | B | CCC | D"),
    ("rating_agencia_rf",   18, False, "SP | MOODYS | FITCH | AUSTIN | SR_RATING | LF_RATING"),
    ("garantias_rf",        30, False, "Descrição das garantias (ex: FGC até R$250k, Alienação fiduciária)"),
]

# Linhas de exemplo
EXEMPLOS = [
    # Ação
    ("PETR4", "Petrobras PN", 100, 28.50, 3250.00, "ACAO", "Energia", "Petrobras",
     None, None, None, None, None, None, None, None, None, None),
    # FII
    ("KNRI11", "Kinea Renda Imobiliária FII", 50, 140.00, 7200.00, "FII", "Imóveis", "Kinea",
     None, None, None, None, None, None, None, None, None, None),
    # ETF
    ("BOVA11", "iShares Ibovespa ETF", 30, 115.00, 3480.00, "ETF", "Diversificado", "BlackRock",
     None, None, None, None, None, None, None, None, None, None),
    # CDB RF
    ("CDB_ITAU_2026", "CDB Itaú 110% CDI", 1, 10000.00, 10850.00, "RENDA_FIXA", "Financeiro", "Itaú Unibanco",
     "CDB", "CDI_PERCENTUAL", 110.0, "31/12/2026", None, "NO_VENCIMENTO", "60872504000123", "AA_MAIS", "SP", "FGC até R$ 250.000"),
    # Tesouro IPCA+
    ("TESOURO_IPCA_2029", "Tesouro IPCA+ 2029", 1, 3500.00, 3780.00, "RENDA_FIXA", "Governo", "Tesouro Nacional",
     "TESOURO_IPCA", "IPCA_MAIS", 5.25, "15/05/2029", None, "NO_VENCIMENTO", "00394460014810", None, None, "Soberano"),
]


def _criar_sheet_carteira(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Carteira"

    cor_obrig = "1F4E79"   # azul escuro
    cor_opcio = "2E75B6"   # azul médio
    cor_rf    = "833C00"   # laranja escuro (RF)
    cor_header_bg = "D6E4F0"

    # ── Cabeçalhos ───────────────────────────────────────────────────────────
    for col_idx, (nome, largura, obrig, _desc) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = Font(bold=True, color="FFFFFF")
        if nome.endswith("_rf"):
            bg = cor_rf
        elif obrig:
            bg = cor_obrig
        else:
            bg = cor_opcio
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Exemplos ─────────────────────────────────────────────────────────────
    for row_idx, exemplo in enumerate(EXEMPLOS, start=2):
        for col_idx, valor in enumerate(exemplo, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal="left")
            # Linhas de exemplo em fundo claro
            cell.fill = PatternFill(fill_type="solid", fgColor="EBF3FB")

    # Linha em branco após os exemplos (pronta para preenchimento)
    ws.cell(row=len(EXEMPLOS) + 2, column=1, value=None)


def _criar_sheet_instrucoes(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Instruções")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80

    linhas = [
        ("COMO USAR", ""),
        ("", ""),
        ("1. Sheet", "Preencha seus dados na aba 'Carteira', a partir da linha 2."),
        ("2. Linhas de exemplo", "As linhas coloridas são exemplos — apague-as antes de enviar."),
        ("3. Colunas obrigatórias", "ticker, nome, quantidade, preco_medio, valor_atual, classe_ativo"),
        ("4. Renda Fixa (RF)", "Preencha os campos subtipo_rf, indexador_rf etc. apenas para ativos de RF."),
        ("5. Datas", "Use o formato DD/MM/AAAA (ex: 31/12/2027) nas colunas de data."),
        ("6. Números", "Use ponto ou vírgula como decimal. Sem separador de milhar."),
        ("7. Upload", "Salve como .xlsx e envie via POST /api/v1/carteira/upload-excel"),
        ("", ""),
        ("CLASSES DE ATIVO VÁLIDAS", ""),
        ("ACAO", "Ações ON, PN, Units negociadas na B3"),
        ("FII", "Fundos de Investimento Imobiliário"),
        ("ETF", "ETFs nacionais e internacionais via B3/BDR"),
        ("RENDA_FIXA", "CDB, LCI, LCA, CRI, CRA, Debêntures, Tesouro Direto"),
        ("BDR", "Brazilian Depositary Receipts"),
        ("FUNDO", "Fundos de Investimento (multimercado, ações, etc.)"),
        ("CRIPTO", "Criptomoedas (Bitcoin, Ethereum, etc.)"),
    ]

    for i, (col_a, col_b) in enumerate(linhas, start=1):
        ca = ws.cell(row=i, column=1, value=col_a)
        cb = ws.cell(row=i, column=2, value=col_b)
        if col_a and col_b == "":
            ca.font = Font(bold=True, size=12, color="1F4E79")
        elif col_a and col_b:
            ca.font = Font(bold=True)


def _criar_sheet_valores_validos(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Valores Válidos")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    dados = [
        ("subtipo_rf", "CDB, LCI, LCA, LIG, CRI, CRA, DEBENTURE, DEBENTURE_INCENTIVADA, "
                       "TESOURO_SELIC, TESOURO_IPCA, TESOURO_PREFIXADO, TESOURO_RENDA_PLUS, "
                       "TESOURO_EDUCA_MAIS, POUPANCA, LC, LF, FIDC, FIAGRO, CPR, CDA_WA, OUTRO"),
        ("indexador_rf", "CDI_PERCENTUAL, CDI_MAIS, SELIC, IPCA_MAIS, PREFIXADO, IGPM_MAIS, TR_MAIS, OUTRO"),
        ("liquidez_rf", "DIARIA, NO_VENCIMENTO, D_MAIS_30, D_MAIS_60, D_MAIS_90"),
        ("rating_escala_rf", "AAA, AA_MAIS, AA, AA_MENOS, A_MAIS, A, A_MENOS, BBB_MAIS, BBB, BBB_MENOS, BB, B, CCC, D"),
        ("rating_agencia_rf", "SP, MOODYS, FITCH, AUSTIN, SR_RATING, LF_RATING"),
        ("classe_ativo", "ACAO, FII, ETF, RENDA_FIXA, BDR, FUNDO, CRIPTO"),
    ]

    ws.cell(row=1, column=1, value="Campo").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Valores aceitos").font = Font(bold=True)

    for i, (campo, valores) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=campo).font = Font(bold=True, color="1F4E79")
        ws.cell(row=i, column=2, value=valores)
